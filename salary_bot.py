"""Telegram-бот обліку заробітку за коробки.

Схема зберігання: один Excel-файл на користувача на місяць — data/{user_id}_{YYYY-MM}.xlsx
Колонки: Дата | Коробки 177 | Коробки 161 | Загальна сума за день
(назви типів коробок задаються константами BIG_LABEL/SMALL_LABEL, можна
перевизначити в config.py, якщо номери коробок знову зміняться)

Запуск:
    python3 salary_bot.py
Токен береться з config.py (BOT_TOKEN) або зі змінної оточення BOT_TOKEN.

Реєстрація:
    При першому /start (якщо ім'я ще не збережено) бота просить людину
    написати ім'я та прізвище — це ім'я потім показується адміну в
    запитах на підтвердження і в списку користувачів. Зберігається в
    data/users.json разом із Telegram-юзернеймом.

Підтвердження адміном:
    Коли звичайний користувач (не з ADMIN_IDS) додає новий день або
    редагує/видаляє вже внесений запис, дані НЕ зберігаються одразу —
    запит стає в чергу. Адміну летить лише легке сповіщення "🔔 Новий
    запит на підтвердження" без деталей і кнопок — щойно один такий
    прийшов, наступні (від того самого чи інших користувачів) вже
    мовчазні, поки адмін не розгребе чергу. Деталі та дії — тільки через
    кнопку меню «⏳ Очікують підтвердження (N)»: список користувачів із
    відкритими запитами → підтвердити/відхилити/відредагувати по черзі.
    Хто з адмінів відповість першим — того рішення і застосовується.
    Користувачу приходить окреме повідомлення з результатом. Записи, які
    вносить сам адмін, зберігаються одразу, без підтвердження. Черга
    запитів зберігається в пам'яті процесу і не переживає перезапуск бота
    (systemctl restart) — це відоме обмеження поточної реалізації.
"""

from __future__ import annotations

import asyncio
import json
import logging
from logging.handlers import RotatingFileHandler
import os
import re
import uuid
from collections import deque
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Iterable, Optional
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from telegram import (
    BotCommand,
    BotCommandScopeChat,
    BotCommandScopeDefault,
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Update,
)
from telegram.error import BadRequest, Conflict, NetworkError
from telegram.ext import (
    Application,
    BaseUpdateProcessor,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)

# --------------------------------------------------------------------------- #
# Конфігурація
# --------------------------------------------------------------------------- #

try:
    import config as _config
except ImportError:
    _config = None


def _setting(name: str, default):
    """config.py -> змінна оточення -> значення за замовчуванням."""
    if _config is not None and hasattr(_config, name):
        return getattr(_config, name)
    env_value = os.getenv(name)
    return env_value if env_value is not None else default


BOT_TOKEN: Optional[str] = _setting("BOT_TOKEN", None)

# Ставки. Виносяться в config.py, якщо треба змінити без правки коду.
RATE_BIG = Decimal(str(_setting("RATE_BIG", "0.90")))
RATE_SMALL = Decimal(str(_setting("RATE_SMALL", "0.70")))
SPLIT = Decimal(str(_setting("SPLIT", "2")))  # ділимо заробіток навпіл

# --------------------------------------------------------------------------- #
# Категорії виробу
#
# На виробництві три групи виробів (коробки, winkel, dekel), кожна у двох
# розмірах — 177 і 161. Людина за день може робити кілька категорій поспіль
# (почала з коробок, перейшла на winkel), тому день зберігає лічильник по
# кожній категорії окремо, а не дві цифри, як було раніше.
#
# Ставки беруться з config.py; за замовчуванням winkel і dekel рахуються за
# тими самими ставками, що й коробки — ПОСТАВ СВОЇ, якщо вони інші:
#     RATE_WINKEL_177 = "1.10"
# Додати ще одну категорію = дописати рядок сюди; формат файлів, звіти та
# екрани вводу підлаштуються самі.
# --------------------------------------------------------------------------- #

CATEGORY_DEFS = [
    ("box_177", "Коробки 177", "RATE_BOX_177", RATE_BIG),
    ("box_161", "Коробки 161", "RATE_BOX_161", RATE_SMALL),
    ("winkel_177", "Winkel 177", "RATE_WINKEL_177", RATE_BIG),
    ("winkel_161", "Winkel 161", "RATE_WINKEL_161", RATE_SMALL),
    ("dekel_177", "Dekel 177", "RATE_DEKEL_177", RATE_BIG),
    ("dekel_161", "Dekel 161", "RATE_DEKEL_161", RATE_SMALL),
]

CATEGORY_KEYS = [key for key, _label, _setting_name, _default in CATEGORY_DEFS]
CATEGORY_LABELS = {
    key: str(_setting(f"LABEL_{key.upper()}", label))
    for key, label, _setting_name, _default in CATEGORY_DEFS
}
CATEGORY_RATES = {
    key: Decimal(str(_setting(setting_name, str(default))))
    for key, _label, setting_name, default in CATEGORY_DEFS
}
CATEGORY_COUNT = len(CATEGORY_KEYS)
EMPTY_COUNTS = tuple(0 for _ in CATEGORY_KEYS)

# Старі назви лишаються для сумісності зі старими файлами й повідомленнями.
BIG_LABEL = CATEGORY_LABELS["box_177"]
SMALL_LABEL = CATEGORY_LABELS["box_161"]

_raw_admins = _setting("ADMIN_IDS", "442336138")
if isinstance(_raw_admins, (set, list, tuple)):
    ADMIN_IDS = {int(x) for x in _raw_admins}
else:
    ADMIN_IDS = {int(x) for x in re.findall(r"\d+", str(_raw_admins))}

BERLIN_TZ = ZoneInfo(str(_setting("TIMEZONE", "Europe/Berlin")))

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)

LOG_FILE = DATA_DIR / "salary_bot.log"
USERS_FILE = DATA_DIR / "users.json"

SHEET_NAME = "Salary"
COLUMNS = ["Дата"] + [CATEGORY_LABELS[key] for key in CATEGORY_KEYS] + ["Загальна сума за день"]
TOTAL_INDEX = len(COLUMNS) - 1  # остання колонка — сума за день

# Заголовки старих файлів -> ключ категорії. Потрібно, щоб файли, створені до
# появи winkel/dekel, читалися без ручної міграції.
LEGACY_HEADER_MAP = {
    "коробки 177": "box_177",
    "коробки 161": "box_161",
    "великі коробки": "box_177",
    "малі коробки": "box_161",
}

TELEGRAM_TEXT_LIMIT = 3500  # запас до ліміту 4096

# --------------------------------------------------------------------------- #
# Логування
# --------------------------------------------------------------------------- #

# Лог читає людина, а не машина: без мілісекунд, з вирівняним рівнем і з
# іменем користувача поруч з його id (див. user_tag).
# Позначка збірки: видно в першому рядку лога після старту. Якщо після заміни
# файлу дата тут стара — значить, працює старий процес і бот не перезапустився.
BOT_VERSION = "2026-08-21 · категорії, доступ, кнопки «Назад»"

LOG_FORMAT = "%(asctime)s %(levelname)-5s %(message)s"
LOG_DATE_FORMAT = "%d.%m.%Y %H:%M:%S"

logger = logging.getLogger("salary_bot")
if not logger.handlers:
    logger.setLevel(logging.INFO)
    # Ротація: до 5 МБ на файл, 5 старих копій (salary_bot.log.1 … .5). Для
    # команди 20-30 людей цього з запасом вистачає на місяці історії, а без
    # ротації лог ріс би необмежено роками роботи сервісу.
    _handler = RotatingFileHandler(LOG_FILE, maxBytes=5 * 1024 * 1024, backupCount=5, encoding="utf-8")
    _handler.setFormatter(logging.Formatter(LOG_FORMAT, datefmt=LOG_DATE_FORMAT))
    logger.addHandler(_handler)
    _console = logging.StreamHandler()
    _console.setFormatter(logging.Formatter(LOG_FORMAT, datefmt=LOG_DATE_FORMAT))
    logger.addHandler(_console)

# --------------------------------------------------------------------------- #
# Стани діалогу та шаблони
# --------------------------------------------------------------------------- #

(
    WAIT_DATE,
    WAIT_CATEGORIES,   # екран з галочками «що робив цього дня»
    WAIT_COUNT,        # кількість по поточній категорії
    WAIT_CONFIRM,
    WAIT_NEXT_ACTION,
    WAIT_MONTH_INPUT,
    WAIT_REGISTER_NAME,
) = range(7)

MONTH_PATTERN = re.compile(r"^\d{4}-\d{2}$")
DATE_PATTERN = re.compile(r"^(\d{1,2})[-.,](\d{1,2})[-.,](\d{4})$")
USER_FILE_PATTERN = re.compile(r"^(\d+)_(\d{4}-\d{2})\.xlsx$")
DATE_CELL_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}$")

RE_MONTH_TOTAL = re.compile(r"^month_total_(\d{4}-\d{2})$")
RE_DOWNLOAD = re.compile(r"^download_(\d{4}-\d{2})$")
RE_ADMIN_USER = re.compile(r"^admin_user_(\d+)$")
RE_ADMIN_LOG_USER = re.compile(r"^admin_log_user_(\d+)$")
RE_ADMIN_SUMMARY = re.compile(r"^admin_summary_(\d{4}-\d{2})$")

# Запити на підтвердження запису адміном (review_<дія>_<id>, id — 8 hex символів).
RE_REVIEW_APPROVE = re.compile(r"^review_approve_([0-9a-f]{8})$")
RE_REVIEW_REJECT = re.compile(r"^review_reject_([0-9a-f]{8})$")
RE_REVIEW_EDIT = re.compile(r"^review_edit_([0-9a-f]{8})$")
RE_REVIEW_USER = re.compile(r"^review_user_(\d+)$")
RE_ACCESS_OK = re.compile(r"^access_ok_(\d+)$")
RE_ACCESS_NO = re.compile(r"^access_no_(\d+)$")

CALLBACK_PATTERN = re.compile(
    r"^(add_more|show_stats|show_month_total|edit_day_menu|close_entry"
    r"|back_to_date|back_to_categories"
    r"|month_total_\d{4}-\d{2}|download_\d{4}-\d{2}"
    r"|admin_users|admin_user_\d+|admin_logs|admin_log_by_user|admin_log_file"
    r"|admin_log_user_\d+"
    r"|admin_monthly_summary|admin_summary_\d{4}-\d{2}"
    r"|review_approve_[0-9a-f]{8}|review_reject_[0-9a-f]{8}|review_edit_[0-9a-f]{8}"
    r"|review_user_\d+"
    r"|pending_reviews|my_pending_reviews"
    r"|access_requests|access_ok_\d+|access_no_\d+|noop"
    r"|date_today)$"
)

# --------------------------------------------------------------------------- #
# Гроші та дати
# --------------------------------------------------------------------------- #


def money(value) -> Decimal:
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def format_money(value) -> str:
    return f"{money(value):.2f}"


def normalize_counts(counts) -> tuple[int, ...]:
    """Будь-що (dict, список, кортеж) -> кортеж лічильників у порядку CATEGORY_KEYS."""
    if isinstance(counts, dict):
        return tuple(max(0, int(counts.get(key, 0) or 0)) for key in CATEGORY_KEYS)
    values = list(counts or ())
    values += [0] * (CATEGORY_COUNT - len(values))
    return tuple(max(0, int(value or 0)) for value in values[:CATEGORY_COUNT])


def category_money(key: str, count: int) -> Decimal:
    return money(Decimal(int(count)) * CATEGORY_RATES[key] / SPLIT)


def calc_day_total(counts) -> tuple[dict[str, Decimal], Decimal]:
    """Повертає (сума по кожній категорії, разом за день)."""
    values = normalize_counts(counts)
    per_category = {key: category_money(key, count) for key, count in zip(CATEGORY_KEYS, values)}
    return per_category, money(sum(per_category.values(), Decimal("0.00")))


def day_total(counts) -> Decimal:
    return calc_day_total(counts)[1]


def counts_summary(counts, only_filled: bool = True) -> str:
    """Рядок «Коробки 177: 40, Winkel 161: 12» для повідомлень."""
    values = normalize_counts(counts)
    parts = [
        f"{CATEGORY_LABELS[key]}: {count}"
        for key, count in zip(CATEGORY_KEYS, values)
        if count or not only_filled
    ]
    return ", ".join(parts) if parts else "—"


def counts_lines(counts, only_filled: bool = True) -> list[str]:
    """Те саме, але окремими рядками — для екранів перевірки та звітів."""
    values = normalize_counts(counts)
    per_category, _ = calc_day_total(values)
    return [
        f"{CATEGORY_LABELS[key]}: {count} × {CATEGORY_RATES[key]} / {SPLIT} = {format_money(per_category[key])} €"
        for key, count in zip(CATEGORY_KEYS, values)
        if count or not only_filled
    ]


def has_any_count(counts) -> bool:
    return any(normalize_counts(counts))


def parse_date_input(raw_date: str) -> Optional[datetime]:
    normalized = raw_date.strip().replace(".", "-").replace(",", "-")
    try:
        return datetime.strptime(normalized, "%d-%m-%Y")
    except ValueError:
        return None


def today_in_berlin():
    return datetime.now(BERLIN_TZ).date()


def current_month() -> str:
    return datetime.now(BERLIN_TZ).strftime("%Y-%m")


# --------------------------------------------------------------------------- #
# Шар зберігання (синхронні функції — викликаються через asyncio.to_thread)
# --------------------------------------------------------------------------- #

_locks: dict[str, asyncio.Lock] = {}


def _lock_for(key: str) -> asyncio.Lock:
    """Один Lock на файл — щоб паралельні записи не губили дані."""
    lock = _locks.get(key)
    if lock is None:
        lock = asyncio.Lock()
        _locks[key] = lock
    return lock


def user_file(user_id: int, month: str) -> Path:
    return DATA_DIR / f"{user_id}_{month}.xlsx"


def _save_workbook(workbook, path: Path) -> None:
    """Атомарне збереження книги: пишемо в сусідній тимчасовий файл і лише
    потім підміняємо ним бойовий через os.replace (на одній ФС це атомарна
    операція). Інакше падіння процесу чи брак місця посеред workbook.save()
    лишили б на диску напівзаписаний, нечитабельний файл — тобто втрату
    місячних даних людини."""
    tmp_path = path.with_name(f".{path.stem}.tmp{path.suffix}")
    try:
        workbook.save(tmp_path)
        os.replace(tmp_path, path)
    except BaseException:
        try:
            tmp_path.unlink()
        except OSError:
            pass
        raise
    finally:
        _invalidate_month_cache(path)


def _ensure_workbook(path: Path) -> None:
    if path.exists():
        return
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(COLUMNS)
    _save_workbook(workbook, path)
    logger.info("Створено новий файл місяця: %s", path.name)


def _as_int(value) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _row_layout(header_row) -> tuple[dict[str, int], int]:
    """За рядком заголовків визначає, у якій колонці яка категорія і де сума.

    Потрібно для файлів, створених до появи winkel/dekel (і ще старіших, на 6
    колонок): за заголовками зрозуміло, що «Коробки 177» тепер лежать у першій
    колонці категорій, а решта категорій у тому файлі просто нульові.
    """
    positions: dict[str, int] = {}
    total_index = None
    labels_to_key = {label.strip().lower(): key for key, label in CATEGORY_LABELS.items()}
    labels_to_key.update(LEGACY_HEADER_MAP)

    for index, cell in enumerate(header_row or ()):
        if cell is None:
            continue
        name = str(cell).strip().lower()
        if not name:
            continue
        if name.startswith("загальна сума") or name.startswith("сума"):
            total_index = index
            continue
        key = labels_to_key.get(name)
        if key is not None and key not in positions:
            positions[key] = index

    if not positions:
        # Заголовка немає або він нечитабельний — вважаємо це найдавнішим
        # форматом: дата | великі | (сміття) | малі | (сміття) | сума.
        return {"box_177": 1, "box_161": 3}, 5
    return positions, total_index


def _row_to_record(row, positions: dict[str, int], total_index: Optional[int]):
    """Рядок аркуша -> (дата, лічильники, сума) або None, якщо це не запис дня."""
    if not row or row[0] is None:
        return None
    date_value = str(row[0]).strip()
    if not DATE_CELL_PATTERN.match(date_value):
        return None  # заголовок або сміття

    counts = tuple(
        _as_int(row[positions[key]]) if key in positions and len(row) > positions[key] else 0
        for key in CATEGORY_KEYS
    )
    stored = row[total_index] if total_index is not None and len(row) > total_index else None
    try:
        total = money(stored) if stored is not None else None
    except (InvalidOperation, TypeError, ValueError):
        total = None
    if total is None:
        total = day_total(counts)
    return date_value, counts, total


def _open_sheet(path: Path):
    """Відкриває аркуш і за потреби приводить його до поточного набору колонок.

    Обробляє:
      * файли до появи winkel/dekel (лише коробки) — нові категорії стають 0;
      * найдавніший формат на 6 колонок;
      * пошкоджені файли (порожні рядки, дубльований заголовок посеред даних).
    Аркуш перебудовується з нуля — правити наявний через delete_cols не можна,
    бо openpyxl не скидає внутрішній лічильник рядків і дані з\'їжджають униз.
    """
    workbook = load_workbook(path)
    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
    sheet.title = SHEET_NAME

    header = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())]
    positions, total_index = _row_layout(header)
    outdated = positions != {key: index + 1 for index, key in enumerate(CATEGORY_KEYS)} or total_index != TOTAL_INDEX

    records: dict[str, list] = {}
    for row in sheet.iter_rows(min_row=1, values_only=True):
        record = _row_to_record(row, positions, total_index)
        if record is None:
            continue
        date_value, counts, total = record
        records[date_value] = [date_value, *counts, float(total)]  # дублікати: лишається останній

    rows = [records[date_value] for date_value in sorted(records)]
    is_clean = (
        not outdated
        and str(sheet.cell(row=1, column=1).value) == COLUMNS[0]
        and sheet.max_row == len(rows) + 1
    )
    if is_clean:
        return workbook, sheet, False

    workbook.remove(sheet)
    new_sheet = workbook.create_sheet(SHEET_NAME, 0)
    new_sheet.append(COLUMNS)
    for row in rows:
        new_sheet.append(row)
    logger.info(
        "Normalized sheet %s (%s, %d rows)",
        path.name, "old column layout" if outdated else "damaged layout", len(rows),
    )
    return workbook, new_sheet, True


def _rows_from_sheet(sheet) -> list[tuple[str, tuple[int, ...], Decimal]]:
    """Рядки з уже відкритого (нормалізованого) аркуша."""
    positions = {key: index + 1 for index, key in enumerate(CATEGORY_KEYS)}
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = _row_to_record(row, positions, TOTAL_INDEX)
        if record is not None:
            rows.append(record)
    return rows


def _save_day_sync(path: Path, date_value: str, counts) -> Decimal:
    _ensure_workbook(path)
    workbook, sheet, _ = _open_sheet(path)
    values = normalize_counts(counts)
    total = day_total(values)

    updated = False
    for row_idx in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row_idx, column=1).value) == date_value:
            for offset, count in enumerate(values):
                sheet.cell(row=row_idx, column=2 + offset, value=int(count))
            sheet.cell(row=row_idx, column=TOTAL_INDEX + 1, value=float(total))
            updated = True
            break
    if not updated:
        sheet.append([date_value, *(int(count) for count in values), float(total)])

    _save_workbook(workbook, path)

    # Новий стан місяця вже є в пам\'яті (та сама книга, яку щойно записали) —
    # кладемо його в кеш, щоб «Разом за місяць» одразу після збереження не
    # відкривав файл вдруге.
    _cache_rows_after_write(path, _rows_from_sheet(sheet))
    return total


def _delete_day_sync(path: Path, date_value: str) -> tuple[bool, int]:
    """Видаляє запис за дату. Повертає (чи видалено, скільки днів лишилось).

    Якщо після видалення в місяці не лишається жодного дня — файл видаляється,
    щоб порожній місяць не з'являвся в меню та в списках.
    """
    if not path.exists():
        return False, 0

    workbook, sheet, changed = _open_sheet(path)
    rows = [
        list(row)[:len(COLUMNS)]
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row and row[0] is not None and DATE_CELL_PATTERN.match(str(row[0]).strip())
    ]
    kept = [row for row in rows if str(row[0]).strip() != date_value]

    if len(kept) == len(rows):
        if changed:
            _save_workbook(workbook, path)
        return False, len(kept)

    if not kept:
        workbook.close()
        path.unlink()
        _invalidate_month_cache(path)
        logger.info("Прибрано порожній файл місяця: %s", path.name)
        return True, 0

    workbook.remove(sheet)
    new_sheet = workbook.create_sheet(SHEET_NAME, 0)
    new_sheet.append(COLUMNS)
    for row in kept:
        new_sheet.append(row)
    _save_workbook(workbook, path)
    _cache_rows_after_write(path, _rows_from_sheet(new_sheet))
    return True, len(kept)


def _read_day_sync(path: Path, date_value: str) -> Optional[tuple[int, ...]]:
    for row_date, counts, _total in _month_rows_sync(path):
        if row_date == date_value:
            return counts
    return None


def _month_statistics_sync(path: Path) -> dict:
    """Підсумок місяця для екрана «Статистика»: заробіток, робочі дні, середнє
    за день і скільки чого зроблено. Рахується з тих самих кешованих рядків,
    що й усе інше, тож зайвого читання файлу немає."""
    rows = _month_rows_sync(path)
    worked = [row for row in rows if any(row[1])]
    total = money(sum((row[2] for row in worked), Decimal("0.00")))
    counts = [0] * CATEGORY_COUNT
    best_day, best_money = None, Decimal("0.00")
    for date_value, day_counts, day_money in worked:
        for index, count in enumerate(day_counts):
            counts[index] += count
        if day_money > best_money:
            best_day, best_money = date_value, day_money
    days = len(worked)
    return {
        "total": total,
        "days": days,
        "average": money(total / days) if days else Decimal("0.00"),
        "counts": tuple(counts),
        "best_day": best_day,
        "best_money": best_money,
    }


def _read_total_sync(path: Path) -> Decimal:
    return money(sum((row[2] for row in _month_rows_sync(path)), Decimal("0.00")))


# --------------------------------------------------------------------------- #
# Кеш розібраних місяців
#
# Раніше кожне читання (сума за місяць, перевірка «чи є дані», звіт, зведення)
# відкривало книгу заново через _open_sheet — а це повний розбір + інколи
# перезапис файлу. При 30 людях і зведенні за місяць виходили десятки відкриттів
# книг поспіль, і бот «задумувався». Тепер розібрані рядки лежать у пам'яті й
# перечитуються лише коли змінився сам файл (звіряємось по mtime+size).
# --------------------------------------------------------------------------- #

# ім'я файлу -> (mtime, size, [(дата, лічильники по категоріях, сума за день)])
_rows_cache: dict[str, tuple[float, int, list[tuple[str, tuple[int, ...], Decimal]]]] = {}


def _invalidate_month_cache(path: Path) -> None:
    _rows_cache.pop(path.name, None)


def _parse_month_file(path: Path) -> list[tuple[str, tuple[int, ...], Decimal]]:
    """Читає файл у режимі read_only (без побудови повної моделі книги) і
    повертає рядки, відсортовані за датою. Розуміє і старі формати (лише
    коробки, 6 колонок) — переписувати файл на диску тут не треба, це зробить
    найближчий запис через _open_sheet."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        positions, total_index = _row_layout(header)

        records: dict[str, tuple[tuple[int, ...], Decimal]] = {}
        for row in sheet.iter_rows(min_row=1, values_only=True):
            record = _row_to_record(row, positions, total_index)
            if record is None:
                continue
            date_value, counts, total = record
            records[date_value] = (counts, total)  # дублікати: лишається останній
    finally:
        workbook.close()

    return [(date_value, counts, total) for date_value, (counts, total) in sorted(records.items())]


def _month_rows_sync(path: Path) -> list[tuple[str, tuple[int, ...], Decimal]]:
    """Рядки місяця з кешу; книга відкривається лише якщо файл змінився."""
    try:
        stat = path.stat()
    except OSError:
        _rows_cache.pop(path.name, None)
        return []

    cached = _rows_cache.get(path.name)
    if cached is not None and cached[0] == stat.st_mtime and cached[1] == stat.st_size:
        return cached[2]

    try:
        rows = _parse_month_file(path)
    except Exception as error:  # noqa: BLE001 — пошкоджений файл не має валити бота
        logger.warning("Cannot read %s: %s", path.name, error)
        return []

    _rows_cache[path.name] = (stat.st_mtime, stat.st_size, rows)
    return rows


def _cache_rows_after_write(path: Path, rows: list[tuple[str, tuple[int, ...], Decimal]]) -> None:
    """Після запису кладемо в кеш уже відомий результат — щоб наступний
    «Разом за місяць» одразу після збереження не відкривав книгу вдруге."""
    try:
        stat = path.stat()
    except OSError:
        _rows_cache.pop(path.name, None)
        return
    _rows_cache[path.name] = (stat.st_mtime, stat.st_size, sorted(rows))


def _month_stats_sync(path: Path) -> tuple[int, bool]:
    """Повертає (кількість рядків з датами, чи є хоч один ненульовий день)."""
    rows = _month_rows_sync(path)
    return len(rows), any(any(counts) for _date, counts, _total in rows)


def _month_has_data_sync(path: Path) -> bool:
    return _month_stats_sync(path)[1]


def cleanup_empty_files() -> int:
    """Прибирає файли без жодного рядка даних (лише заголовок). Викликається при старті."""
    removed = 0
    for path in sorted(DATA_DIR.glob("*.xlsx")):
        if not USER_FILE_PATTERN.match(path.name):
            continue
        if _month_stats_sync(path)[0] > 0:
            continue
        try:
            path.unlink()
            _invalidate_month_cache(path)
            removed += 1
            logger.info("Старт: прибрано порожній файл місяця %s", path.name)
        except OSError as error:
            logger.warning("Cannot remove %s: %s", path.name, error)
    return removed


def _list_user_files_sync(
    month: Optional[str] = None,
    user_id: Optional[int] = None,
    require_data: bool = True,
) -> list[Path]:
    result = []
    for path in sorted(DATA_DIR.glob("*.xlsx")):
        match = USER_FILE_PATTERN.match(path.name)
        if not match:
            continue
        if user_id is not None and int(match.group(1)) != user_id:
            continue
        if month is not None and match.group(2) != month:
            continue
        if require_data and not _month_has_data_sync(path):
            continue
        result.append(path)
    return result


def _list_months_sync(user_id: Optional[int] = None) -> list[str]:
    months = set()
    for path in _list_user_files_sync(user_id=user_id):
        match = USER_FILE_PATTERN.match(path.name)
        if match:
            months.add(match.group(2))
    return sorted(months)


def _list_user_ids_sync() -> list[int]:
    ids = set()
    for path in _list_user_files_sync():
        match = USER_FILE_PATTERN.match(path.name)
        if match:
            ids.add(int(match.group(1)))
    return sorted(ids)


def _read_all_totals_sync(month: str) -> tuple[Decimal, list[str]]:
    total = Decimal("0.00")
    names = []
    for path in _list_user_files_sync(month=month):
        total += _read_total_sync(path)
        names.append(path.name)
    return money(total), names


def _month_snapshot_sync(month: str) -> tuple[list[dict], list[str], dict[tuple[str, int], tuple[tuple[int, ...], Decimal]]]:
    """Один прохід по файлах місяця для обох аркушів зведення.

    Раніше «Зведення» і «Деталізація за днями» рахувалися двома окремими
    функціями, кожна з яких відкривала всі файли місяця — при 30 людях це 60
    відкриттів книг на одне натискання кнопки. Тепер файли читаються один раз
    (та ще й з кешу), а далі дані просто розкладаються по двох структурах:
      * підсумок на кожного користувача (відсортований за заробітком);
      * дати місяця + мапа (дата, user_id) -> (лічильники, сума за день).
    """
    user_map = _load_user_map_sync()
    summary: list[dict] = []
    users: list[dict] = []
    cells: dict[tuple[str, int], tuple[tuple[int, ...], Decimal]] = {}
    dates: set[str] = set()

    for path in _list_user_files_sync(month=month):
        match = USER_FILE_PATTERN.match(path.name)
        if not match:
            continue
        uid = int(match.group(1))

        entry = user_map.get(str(uid), {})
        name = (entry.get("registered_name") or entry.get("telegram_label") or "").strip() or str(uid)
        users.append({"user_id": uid, "name": name})

        totals = [0] * CATEGORY_COUNT
        money_sum = Decimal("0.00")
        for date_value, counts, day_money in _month_rows_sync(path):
            for index, count in enumerate(counts):
                totals[index] += count
            money_sum += day_money
            dates.add(date_value)
            cells[(date_value, uid)] = (counts, day_money)

        summary.append(
            {
                "user_id": uid,
                "name": name,
                "counts": tuple(totals),
                "total_items": sum(totals),
                "money": money(money_sum),
            }
        )

    # У "Зведенні" — хто заробив більше, той згори (за рівних сум — за алфавітом).
    summary.sort(key=lambda r: (-r["money"], r["name"].lower()))
    users.sort(key=lambda u: u["name"].lower())
    return summary, sorted(dates), cells


_GRANDTOTAL_FILL = PatternFill(start_color="FFD6E4F5", end_color="FFD6E4F5", fill_type="solid")


def _write_daily_detail_sheet(
    workbook: Workbook,
    dates: list[str],
    users: list[dict],
    cells: dict[tuple[str, int], tuple[tuple[int, ...], Decimal]],
) -> None:
    """Деталізація за днями: для кожної дати — рядок на кожного користувача, який
    того дня щось вносив (штуки по кожній категорії та сума за день). Комірка
    «Дата» об\'єднана по всьому блоку дня. Внизу — підсумок за місяць."""
    sheet = workbook.create_sheet("Деталізація за днями")
    headers = ["Дата", "Користувач"]
    headers += [f"{CATEGORY_LABELS[key]}, шт" for key in CATEGORY_KEYS]
    headers += ["Разом шт", "Сума за день, €"]
    sheet.append(headers)
    sheet.row_dimensions[1].height = 30
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    money_column = len(headers)
    month_counts = [0] * CATEGORY_COUNT
    month_money = Decimal("0.00")
    row_idx = 2

    for date_value in dates:
        date_label = datetime.strptime(date_value, "%Y-%m-%d").strftime("%d-%m-%Y")
        day_users = [u for u in users if (date_value, u["user_id"]) in cells]
        if not day_users:
            continue

        block_start = row_idx
        for user in day_users:
            counts, day_money = cells[(date_value, user["user_id"])]
            sheet.append([date_label, user["name"], *counts, sum(counts), float(day_money)])
            sheet.cell(row=row_idx, column=money_column).number_format = '0.00" €"'
            for index, count in enumerate(counts):
                month_counts[index] += count
            month_money += day_money
            row_idx += 1

        if len(day_users) > 1:
            sheet.merge_cells(start_row=block_start, start_column=1, end_row=row_idx - 1, end_column=1)
            sheet.cell(row=block_start, column=1).alignment = Alignment(vertical="center")

    sheet.append(["Разом за місяць", "", *month_counts, sum(month_counts), float(money(month_money))])
    for cell in sheet[row_idx]:
        cell.font = Font(bold=True)
        cell.fill = _GRANDTOTAL_FILL
    sheet.cell(row=row_idx, column=money_column).number_format = '0.00" €"'

    sheet.freeze_panes = "C2"
    widths = [13, 24] + [13] * CATEGORY_COUNT + [11, 16]
    for col_idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _build_monthly_summary_sync(month: str) -> Optional[Path]:
    """Будує зведену книгу за місяць: підсумок по кожному користувачу плюс
    деталізація по днях (хто скільки 177/161 зробив і на яку суму кожного дня,
    з підсумком за день і за місяць). Зберігає у тимчасовий xlsx у DATA_DIR.
    Повертає None, якщо за місяць немає даних."""
    rows, dates, cells = _month_snapshot_sync(month)
    if not rows:
        return None
    users = sorted(
        ({"user_id": row["user_id"], "name": row["name"]} for row in rows),
        key=lambda u: u["name"].lower(),
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Зведення"

    headers = ["№", "Користувач", "ID"]
    headers += [CATEGORY_LABELS[key] for key in CATEGORY_KEYS]
    headers += ["Разом шт", "Сума, €"]
    sheet.append(headers)
    money_column = len(headers)
    sheet.row_dimensions[1].height = 30
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for idx, row in enumerate(rows, start=1):
        sheet.append(
            [idx, row["name"], row["user_id"], *row["counts"], row["total_items"], float(row["money"])]
        )
        sheet.cell(row=idx + 1, column=money_column).number_format = '0.00" €"'

    totals = [sum(row["counts"][index] for row in rows) for index in range(CATEGORY_COUNT)]
    total_money = money(sum((row["money"] for row in rows), Decimal("0.00")))

    total_row_idx = len(rows) + 2
    sheet.append(["", "Разом за місяць", "", *totals, sum(totals), float(total_money)])
    for cell in sheet[total_row_idx]:
        cell.font = Font(bold=True)
        cell.fill = _GRANDTOTAL_FILL
    sheet.cell(row=total_row_idx, column=money_column).number_format = '0.00" €"'

    sheet.freeze_panes = "A2"
    widths = [4, 24, 12] + [13] * CATEGORY_COUNT + [11, 13]
    for col_idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    if dates and users:
        _write_daily_detail_sheet(workbook, dates, users, cells)

    tmp_path = DATA_DIR / f"summary_{month}.xlsx"
    _save_workbook(workbook, tmp_path)
    return tmp_path


# --------------------------------------------------------------------------- #
# Місячний звіт для користувача (оформлена копія його файлу)
#
# Робочий файл data/{user_id}_{YYYY-MM}.xlsx лишається технічним — 4 колонки без
# оформлення, як його читає решта коду. Людині ж надсилається окремий, красивий
# звіт: день тижня, суми окремо по кожному типу коробок, підсумковий рядок і
# великий «Дохід за місяць» унизу.
# --------------------------------------------------------------------------- #

WEEKDAYS_UA = ["Понеділок", "Вівторок", "Середа", "Четвер", "П'ятниця", "Субота", "Неділя"]
REPORT_MONEY_FORMAT = '#,##0.00" €"'
_REPORT_HEADER_FILL = PatternFill("solid", fgColor="FF1F4E78")
_REPORT_STRIPE_FILL = PatternFill("solid", fgColor="FFF2F6FA")
_REPORT_WEEKEND_FILL = PatternFill("solid", fgColor="FFFDF0E6")
_REPORT_TOTAL_FILL = PatternFill("solid", fgColor="FFDCE6F1")
_REPORT_THIN = Side(style="thin", color="FFBFBFBF")
_REPORT_BORDER = Border(left=_REPORT_THIN, right=_REPORT_THIN, top=_REPORT_THIN, bottom=_REPORT_THIN)


def _build_user_report_sync(path: Path, month: str, display_name: str = "") -> Optional[Path]:
    """Оформлений місячний звіт. Показуються лише ті категорії, які людина
    цього місяця реально робила — інакше половина таблиці була б нулями.
    Повертає шлях до тимчасового файлу або None, якщо за місяць немає днів."""
    rows = _month_rows_sync(path)
    if not rows:
        return None

    used_keys = [
        key
        for index, key in enumerate(CATEGORY_KEYS)
        if any(counts[index] for _date, counts, _total in rows)
    ] or [CATEGORY_KEYS[0]]
    used_indexes = [CATEGORY_KEYS.index(key) for key in used_keys]

    headers = [("Дата", 13), ("День тижня", 14)]
    headers += [(f"{CATEGORY_LABELS[key]}, шт", 15) for key in used_keys]
    headers += [(f"{CATEGORY_LABELS[key]}, €", 15) for key in used_keys]
    headers += [("Разом за день, €", 17)]
    last_col = len(headers)
    last_letter = get_column_letter(last_col)
    first_money_col = 3 + len(used_keys)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Звіт"

    try:
        month_title = datetime.strptime(month, "%Y-%m").strftime("%m.%Y")
    except ValueError:
        month_title = month

    sheet.merge_cells(f"A1:{last_letter}1")
    title = sheet.cell(row=1, column=1, value=f"Заробіток за {month_title}")
    title.font = Font(bold=True, size=15, color="FF1F4E78")
    title.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 26

    subtitle_parts = []
    if display_name:
        subtitle_parts.append(display_name)
    subtitle_parts.append(
        "Ставки: "
        + " · ".join(f"{CATEGORY_LABELS[key]} {format_money(CATEGORY_RATES[key])} €" for key in used_keys)
        + f" · поділ на {SPLIT}"
    )
    sheet.merge_cells(f"A2:{last_letter}2")
    subtitle = sheet.cell(row=2, column=1, value="   |   ".join(subtitle_parts))
    subtitle.font = Font(size=10, italic=True, color="FF595959")
    subtitle.alignment = Alignment(horizontal="center", vertical="center")

    header_row = 4
    for index, (name, width) in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=index, value=name)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = _REPORT_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _REPORT_BORDER
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[header_row].height = 30

    totals_counts = [0] * len(used_keys)
    totals_money = [Decimal("0.00")] * len(used_keys)
    total_all = Decimal("0.00")
    current = header_row

    for offset, (date_value, counts, stored_total) in enumerate(rows):
        current = header_row + 1 + offset
        per_category, computed_total = calc_day_total(counts)
        row_total = stored_total if stored_total is not None else computed_total
        total_all += row_total

        day = datetime.strptime(date_value, "%Y-%m-%d")
        weekday = day.weekday()
        values = [day.strftime("%d.%m.%Y"), WEEKDAYS_UA[weekday]]
        for position, index in enumerate(used_indexes):
            totals_counts[position] += counts[index]
            values.append(counts[index])
        for position, key in enumerate(used_keys):
            totals_money[position] += per_category[key]
            values.append(float(per_category[key]))
        values.append(float(row_total))

        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=current, column=index, value=value)
            cell.border = _REPORT_BORDER
            if index >= first_money_col:
                cell.number_format = REPORT_MONEY_FORMAT
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center")
            if weekday >= 5:
                cell.fill = _REPORT_WEEKEND_FILL
            elif offset % 2:
                cell.fill = _REPORT_STRIPE_FILL
        sheet.cell(row=current, column=last_col).font = Font(bold=True)

    total_row = current + 1
    totals = ["РАЗОМ", f"{len(rows)} дн."]
    totals += totals_counts
    totals += [float(money(value)) for value in totals_money]
    totals += [float(money(total_all))]
    for index, value in enumerate(totals, start=1):
        cell = sheet.cell(row=total_row, column=index, value=value)
        cell.font = Font(bold=True, size=11)
        cell.fill = _REPORT_TOTAL_FILL
        cell.border = Border(
            left=_REPORT_THIN, right=_REPORT_THIN, bottom=_REPORT_THIN,
            top=Side(style="medium", color="FF1F4E78"),
        )
        if index >= first_money_col:
            cell.number_format = REPORT_MONEY_FORMAT
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.alignment = Alignment(horizontal="center")
    sheet.row_dimensions[total_row].height = 22

    summary_row = total_row + 2
    sheet.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=2)
    label = sheet.cell(row=summary_row, column=1, value="Дохід за місяць:")
    label.font = Font(bold=True, size=13)
    label.alignment = Alignment(horizontal="left", vertical="center")
    amount = sheet.cell(row=summary_row, column=3, value=float(money(total_all)))
    amount.font = Font(bold=True, size=13, color="FF1F4E78")
    amount.number_format = REPORT_MONEY_FORMAT
    amount.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[summary_row].height = 24

    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{last_letter}{total_row - 1}"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.print_title_rows = f"{header_row}:{header_row}"

    report_path = DATA_DIR / f"report_{path.stem}.xlsx"
    _save_workbook(workbook, report_path)
    return report_path


# users.json читався заново при кожному звертанні (ім'я автора запиту, список
# користувачів, зведення), тож кешуємо його так само по mtime+size.
_user_map_cache: Optional[tuple[float, int, dict]] = None


def _load_user_map_sync() -> dict:
    """Повертає {user_id_str: {"telegram_label": str, "registered_name": Optional[str]}}.

    Старий формат (значення — просто рядок з юзернеймом) переводиться в нову
    форму на льоту; на диску перепишеться при першому ж збереженні.
    """
    global _user_map_cache
    try:
        stat = USERS_FILE.stat()
    except OSError:
        _user_map_cache = None
        return {}

    if _user_map_cache is not None and _user_map_cache[0] == stat.st_mtime and _user_map_cache[1] == stat.st_size:
        return _user_map_cache[2]

    try:
        with open(USERS_FILE, "r", encoding="utf-8") as fh:
            raw = json.load(fh)
    except (OSError, json.JSONDecodeError):
        logger.warning("users.json is unreadable, treating as empty")
        return {}

    mapping: dict = {}
    for key, value in raw.items():
        if isinstance(value, dict):
            # Зберігаємо всі поля запису (approved, requested_at тощо), а не
            # лише два відомих — інакше нормалізація гасила б прапорець доступу
            # при кожному збереженні імені.
            entry = dict(value)
            entry["telegram_label"] = value.get("telegram_label", "")
            entry["registered_name"] = value.get("registered_name")
            mapping[key] = entry
        else:
            mapping[key] = {"telegram_label": str(value or ""), "registered_name": None}

    _user_map_cache = (stat.st_mtime, stat.st_size, mapping)
    return mapping


def _save_user_map_sync(mapping: dict) -> None:
    global _user_map_cache
    tmp_path = USERS_FILE.with_suffix(".json.tmp")
    with open(tmp_path, "w", encoding="utf-8") as fh:
        json.dump(mapping, fh, ensure_ascii=False, indent=2)
    os.replace(tmp_path, USERS_FILE)  # атомарно, як і з Excel
    _user_map_cache = None


def _register_user_sync(user_id: int, display_name: str) -> None:
    mapping = _load_user_map_sync()
    entry = mapping.get(str(user_id), {"telegram_label": "", "registered_name": None})
    if entry.get("telegram_label") == display_name:
        return
    entry["telegram_label"] = display_name
    mapping[str(user_id)] = entry
    _save_user_map_sync(mapping)


def _save_registered_name_sync(user_id: int, full_name: str) -> None:
    mapping = _load_user_map_sync()
    entry = mapping.get(str(user_id), {"telegram_label": "", "registered_name": None})
    entry["registered_name"] = full_name
    mapping[str(user_id)] = entry
    _save_user_map_sync(mapping)


def _set_access_sync(user_id: int, approved: Optional[bool], requested_at: Optional[float] = None) -> None:
    mapping = _load_user_map_sync()
    entry = dict(mapping.get(str(user_id)) or {"telegram_label": "", "registered_name": None})
    entry["approved"] = approved
    if requested_at is not None:
        entry["requested_at"] = requested_at
    mapping[str(user_id)] = entry
    _save_user_map_sync(mapping)


def _is_approved_sync(user_id: int) -> bool:
    if user_id in ADMIN_IDS:  # адміни завжди мають доступ
        return True
    entry = _load_user_map_sync().get(str(user_id))
    return bool(isinstance(entry, dict) and entry.get("approved") is True)


def _pending_access_sync() -> list[tuple[int, str]]:
    """Хто чекає на допуск: (id, ім'я) у порядку подання заявки."""
    waiting = []
    for key, entry in _load_user_map_sync().items():
        if not str(key).isdigit() or int(key) in ADMIN_IDS:
            continue
        if not isinstance(entry, dict) or entry.get("approved") is not None:
            continue
        if not entry.get("requested_at"):
            continue
        name = (entry.get("registered_name") or entry.get("telegram_label") or "").strip() or str(key)
        waiting.append((int(key), name, float(entry["requested_at"])))
    waiting.sort(key=lambda item: item[2])
    return [(uid, name) for uid, name, _at in waiting]


def _get_registered_name_sync(user_id: int) -> Optional[str]:
    mapping = _load_user_map_sync()
    entry = mapping.get(str(user_id))
    if isinstance(entry, dict):
        return entry.get("registered_name")
    return None


def _tail_log_sync(lines: int, marker: Optional[str] = None) -> list[str]:
    if not LOG_FILE.exists():
        return []
    with open(LOG_FILE, "r", encoding="utf-8", errors="ignore") as fh:
        if marker is None:
            return [line.rstrip("\n") for line in deque(fh, maxlen=lines)]
        matched = (line.rstrip("\n") for line in fh if marker in line)
        return list(deque(matched, maxlen=lines))


def _filter_log_sync(marker: str) -> list[str]:
    if not LOG_FILE.exists():
        return []
    with open(LOG_FILE, "r", encoding="utf-8", errors="ignore") as fh:
        return [line.rstrip("\n") for line in fh if marker in line]


# --------------------------------------------------------------------------- #
# Асинхронні обгортки (блокуючий I/O — в окремому потоці, під локом)
# --------------------------------------------------------------------------- #


async def save_day(user_id: int, month: str, day: int, counts) -> Decimal:
    path = user_file(user_id, month)
    date_value = f"{month}-{day:02d}"
    values = normalize_counts(counts)
    async with _lock_for(path.name):
        total = await asyncio.to_thread(_save_day_sync, path, date_value, values)
    logger.info(
        "%s — зберіг день %s: %s = %s €",
        user_tag(user_id), _human_date(date_value), counts_summary(values), format_money(total),
    )
    return total


async def delete_day(user_id: int, month: str, day: int) -> tuple[bool, int]:
    path = user_file(user_id, month)
    date_value = f"{month}-{day:02d}"
    async with _lock_for(path.name):
        removed, remaining = await asyncio.to_thread(_delete_day_sync, path, date_value)
    if removed:
        logger.info(
            "%s — видалив день %s (днів у місяці лишилось: %s)",
            user_tag(user_id), _human_date(date_value), remaining,
        )
    else:
        logger.info("%s — спроба видалити день %s, а запису немає", user_tag(user_id), _human_date(date_value))
    return removed, remaining


async def read_day(user_id: int, month: str, day: int) -> Optional[tuple[int, ...]]:
    path = user_file(user_id, month)
    async with _lock_for(path.name):
        return await asyncio.to_thread(_read_day_sync, path, f"{month}-{day:02d}")


async def month_has_data(path: Path) -> bool:
    """Перевірка «чи є записи» поза головним циклом: у найгіршому випадку тут
    читається файл, і робити це в event loop не можна — інакше решта команди
    чекає, поки бот розбирає чужу книгу."""
    return await asyncio.to_thread(_month_has_data_sync, path)


async def month_statistics(user_id: int, month: str) -> dict:
    path = user_file(user_id, month)
    async with _lock_for(path.name):
        return await asyncio.to_thread(_month_statistics_sync, path)


async def read_month_total(user_id: int, month: str) -> Decimal:
    path = user_file(user_id, month)
    async with _lock_for(path.name):
        return await asyncio.to_thread(_read_total_sync, path)


async def read_all_totals(month: str) -> tuple[Decimal, list[str]]:
    return await asyncio.to_thread(_read_all_totals_sync, month)


async def build_monthly_summary(month: str) -> Optional[Path]:
    return await asyncio.to_thread(_build_monthly_summary_sync, month)


async def build_user_report(user_id: int, month: str, display_name: str = "") -> Optional[Path]:
    path = user_file(user_id, month)
    async with _lock_for(path.name):
        return await asyncio.to_thread(_build_user_report_sync, path, month, display_name)


async def list_months(user_id: Optional[int]) -> list[str]:
    return await asyncio.to_thread(_list_months_sync, user_id)


async def list_user_ids() -> list[int]:
    return await asyncio.to_thread(_list_user_ids_sync)


async def list_user_files(month: str) -> list[Path]:
    return await asyncio.to_thread(_list_user_files_sync, month, None)


async def load_user_map() -> dict:
    return await asyncio.to_thread(_load_user_map_sync)


_registered_users: dict[int, str] = {}


async def register_user(user) -> None:
    """Записує користувача в users.json. Повторні виклики з тим самим іменем безкоштовні."""
    if user is None or getattr(user, "id", None) is None:
        return
    display_name = getattr(user, "username", None) or getattr(user, "full_name", "") or ""
    if _registered_users.get(user.id) == display_name:
        return
    async with _lock_for(USERS_FILE.name):
        await asyncio.to_thread(_register_user_sync, user.id, display_name)
    _registered_users[user.id] = display_name


async def known_users() -> dict[int, tuple[str, bool]]:
    """Усі відомі користувачі: {id: (ім'я, чи є дані)}.

    Джерела об'єднуються — users.json (усі, хто хоч раз запускав бота)
    і файли даних (на випадок, якщо users.json загубився). Як ім'я
    пріоритетно береться те, яке людина вписала при реєстрації, інакше —
    її Telegram-юзернейм/ім'я, автоматично зафіксовані при /start.
    """
    with_data = set(await list_user_ids())
    user_map = await load_user_map()

    result: dict[int, tuple[str, bool]] = {}
    for key, entry in user_map.items():
        if str(key).isdigit():
            uid = int(key)
            label = (entry.get("registered_name") or entry.get("telegram_label") or "") if isinstance(entry, dict) else ""
            result[uid] = (label, uid in with_data)
    for uid in with_data:
        result.setdefault(uid, ("", True))
    return dict(sorted(result.items()))


async def save_registered_name(user_id: int, full_name: str) -> None:
    async with _lock_for(USERS_FILE.name):
        await asyncio.to_thread(_save_registered_name_sync, user_id, full_name)


_ACTION_WORDS = {"add": "новий запис", "edit": "зміну", "delete": "видалення"}


def _human_date(date_value: str) -> str:
    """2026-08-15 -> 15.08.2026 (у лог пишемо так, як людина бачить у боті)."""
    try:
        return datetime.strptime(date_value, "%Y-%m-%d").strftime("%d.%m.%Y")
    except ValueError:
        return date_value


def user_tag(user_id: int, name: Optional[str] = None) -> str:
    """Мітка користувача для лога: «[USER:442336138] Тарас Хомин».

    Ім'я береться з users.json (він у кеші, тож це не читання з диска). Тег
    "[USER:<id>]" лишився незмінним навмисно — по ньому працює фільтрація
    логів по користувачу в адмін-меню та /logtail.
    """
    if name is None:
        entry = _load_user_map_sync().get(str(user_id))
        if isinstance(entry, dict):
            name = (entry.get("registered_name") or entry.get("telegram_label") or "").strip()
    label = f"[USER:{user_id}]"
    return f"{label} {name}" if name else label


def admin_tag(user_id: int) -> str:
    """Те саме для адміна: «[ADMIN:442336138] Тарас Хомин»."""
    return user_tag(user_id).replace("[USER:", "[ADMIN:", 1)


async def user_allowed(user_id: int) -> bool:
    return await asyncio.to_thread(_is_approved_sync, user_id)


async def set_access(user_id: int, approved: Optional[bool], requested_at: Optional[float] = None) -> None:
    async with _lock_for(USERS_FILE.name):
        await asyncio.to_thread(_set_access_sync, user_id, approved, requested_at)


async def pending_access() -> list[tuple[int, str]]:
    return await asyncio.to_thread(_pending_access_sync)


async def request_access(context: ContextTypes.DEFAULT_TYPE, user_id: int, display_name: str) -> None:
    """Ставить нового користувача в чергу на допуск і будить адмінів. Повторний
    /start заявку не дублює — просто оновлює час."""
    await set_access(user_id, None, requested_at=datetime.now(BERLIN_TZ).timestamp())
    logger.info("%s — просить доступ до бота", user_tag(user_id, display_name))
    text = (
        f"🔐 Новий користувач просить доступ: {display_name} (ID {user_id}).\n"
        "Відкрий «🔐 Заявки на доступ» у меню."
    )
    for admin_id in ADMIN_IDS:
        try:
            await context.bot.send_message(admin_id, text)
        except Exception as error:  # noqa: BLE001 — недоступний адмін не має ламати реєстрацію
            logger.warning("Could not notify admin %s about access request: %s", admin_id, error)


async def get_registered_name(user_id: int) -> Optional[str]:
    return await asyncio.to_thread(_get_registered_name_sync, user_id)


async def tail_log(lines: int, marker: Optional[str] = None) -> list[str]:
    return await asyncio.to_thread(_tail_log_sync, lines, marker)


async def filter_log(marker: str) -> list[str]:
    return await asyncio.to_thread(_filter_log_sync, marker)


# --------------------------------------------------------------------------- #
# Клавіатури та безпечне редагування повідомлень
# --------------------------------------------------------------------------- #


def is_admin(user_id: Optional[int]) -> bool:
    return user_id is not None and user_id in ADMIN_IDS


def build_main_menu(user_id: Optional[int] = None, month: Optional[str] = None) -> InlineKeyboardMarkup:
    month = month or current_month()

    # «🏠 Меню» і «📥 Завантажити файл» з головного меню прибрані: меню й так
    # уже на екрані, а файл віддається після «📊 Сума за місяць».
    buttons = [
        [
            InlineKeyboardButton("➕ Внести дані", callback_data="add_more"),
            InlineKeyboardButton("📊 Статистика", callback_data="show_stats"),
        ],
        [InlineKeyboardButton("📝 Редагувати дані", callback_data="edit_day_menu")],
    ]
    if is_admin(user_id):
        waiting_access = len(_pending_access_sync())
        if waiting_access:
            buttons.append(
                [InlineKeyboardButton(f"🔐 Заявки на доступ ({waiting_access})", callback_data="access_requests")]
            )
        pending_count = _pending_review_count()
        buttons.append(
            [InlineKeyboardButton(f"⏳ Очікують підтвердження ({pending_count})", callback_data="pending_reviews")]
        )
        buttons.append(
            [
                InlineKeyboardButton("👥 Користувачі", callback_data="admin_users"),
                InlineKeyboardButton("🗒️ Логи", callback_data="admin_logs"),
            ]
        )
        buttons.append(
            [InlineKeyboardButton("📈 Зведення за місяць", callback_data="admin_monthly_summary")]
        )
    elif user_id is not None:
        my_pending = _user_pending_count(user_id)
        if my_pending:
            buttons.append(
                [
                    InlineKeyboardButton(
                        f"⏳ Очікують підтвердження ({my_pending})", callback_data="my_pending_reviews"
                    )
                ]
            )
    return InlineKeyboardMarkup(buttons)


def build_date_keyboard(user_id: Optional[int] = None) -> InlineKeyboardMarkup:
    """Екран вводу дати — з\'являється лише після «Внести дані» / «Редагувати».

    Дані вносять переважно того ж вечора, тож набирати «21-08-2026» руками —
    найчастіша й найнудніша дія в боті. Ввід дати текстом нікуди не подівся:
    будь-яку іншу дату так само можна написати.
    """
    today = today_in_berlin()
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton(f"📅 Сьогодні ({today.strftime('%d.%m')})", callback_data="date_today")],
            [InlineKeyboardButton("🏠 Меню", callback_data="close_entry")],
        ]
    )


def build_month_selection_keyboard(months: list[str], back_callback: Optional[str] = None) -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton(month, callback_data=f"month_total_{month}")] for month in months]
    last_row = []
    if back_callback:
        last_row.append(InlineKeyboardButton("⬅️ Назад", callback_data=back_callback))
    last_row.append(InlineKeyboardButton("🏠 Меню", callback_data="close_entry"))
    rows.append(last_row)
    return InlineKeyboardMarkup(rows)


def build_month_result_keyboard(month: str, back_callback: Optional[str] = None) -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton("📥 Завантажити детальний звіт", callback_data=f"download_{month}")]]
    rows.append([InlineKeyboardButton("🗓 Інший місяць", callback_data="show_month_total")])
    last_row = []
    if back_callback:
        last_row.append(InlineKeyboardButton("⬅️ Назад", callback_data=back_callback))
    last_row.append(InlineKeyboardButton("🏠 Меню", callback_data="close_entry"))
    rows.append(last_row)
    return InlineKeyboardMarkup(rows)


def build_admin_summary_month_keyboard(months: list[str]) -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton(month, callback_data=f"admin_summary_{month}")] for month in months]
    rows.append([InlineKeyboardButton("🏠 Меню", callback_data="close_entry")])
    return InlineKeyboardMarkup(rows)


# Одне «активне» меню на чат: message_id останнього повідомлення з робочими
# кнопками. Перед тим як показати нове меню новим повідомленням (не редагуючи
# старе), кнопки з попереднього прибираються — щоб у чаті ніколи не лишалось
# декількох одночасно клікабельних меню.
_active_menu_message: dict[int, int] = {}


async def _clear_previous_menu(bot, chat_id: int) -> None:
    old_message_id = _active_menu_message.get(chat_id)
    if old_message_id is None:
        return
    try:
        await bot.edit_message_reply_markup(chat_id=chat_id, message_id=old_message_id, reply_markup=None)
    except Exception:  # noqa: BLE001 — повідомлення могли вже видалити чи відредагувати вручну
        pass


# Telegram дозволяє відповісти на натискання кнопки рівно один раз. Раніше
# on_callback одразу робив порожній query.answer(), тож усі подальші
# «немає даних» просто не долітали до людини — кнопка блимала й нічого не
# з'являлось. Тепер відповідь одна: або змістовна, або порожня в кінці.
_answered_queries: set[str] = set()
_answered_order: deque[str] = deque(maxlen=1000)


async def ack(query: CallbackQuery, text: Optional[str] = None, alert: bool = False) -> None:
    """Відповідь на натискання кнопки. alert=False — сповіщення, що само
    зникає за кілька секунд; alert=True — вікно з кнопкою «ОК»."""
    query_id = getattr(query, "id", None)
    if query_id is not None:
        if query_id in _answered_queries:
            return
        if len(_answered_order) == _answered_order.maxlen and _answered_order:
            _answered_queries.discard(_answered_order[0])
        _answered_order.append(query_id)
        _answered_queries.add(query_id)
    try:
        await query.answer(text or "", show_alert=alert)
    except BadRequest as error:  # застаріле натискання — не привід падати
        logger.debug("Could not answer callback query: %s", error)


async def safe_edit(query: CallbackQuery, text: str, reply_markup=None) -> None:
    """edit_message_text, який не падає на 'Message is not modified'.

    Редагує те саме повідомлення на місці, тож воно й лишається єдиним
    активним меню в чаті — просто оновлюємо трекер на випадок наступного
    send_menu_reply (щоб було що прибирати)."""
    try:
        await query.edit_message_text(text, reply_markup=reply_markup)
    except BadRequest as error:
        if "Message is not modified" not in str(error):
            raise
        logger.debug("Ignored 'Message is not modified'")

    chat_id = query.message.chat_id
    if reply_markup is not None:
        _active_menu_message[chat_id] = query.message.message_id
    else:
        _active_menu_message.pop(chat_id, None)


async def send_menu_reply(update: Update, context: ContextTypes.DEFAULT_TYPE, text: str, reply_markup=None):
    """Заміна update.message.reply_text для «менюшних» відповідей: перед тим як
    надіслати нове повідомлення, прибирає кнопки з попереднього активного меню
    в цьому чаті, щоб завжди залишалось рівно одне робоче меню, а не купа
    старих клікабельних вікон одне під одним."""
    chat_id = update.effective_chat.id
    await _clear_previous_menu(context.bot, chat_id)
    message = await update.message.reply_text(text, reply_markup=reply_markup)
    if reply_markup is not None:
        _active_menu_message[chat_id] = message.message_id
    else:
        _active_menu_message.pop(chat_id, None)
    return message


def _lines_word(count: int) -> str:
    if count % 10 == 1 and count % 100 != 11:
        return "рядок"
    if 2 <= count % 10 <= 4 and not (12 <= count % 100 <= 14):
        return "рядки"
    return "рядків"


def _safe_filename(text: str) -> str:
    """Ім'я для файлу: без пробілів і символів, які Telegram чи ФС не люблять."""
    cleaned = re.sub(r"[^\w\-.]+", "_", text.strip(), flags=re.UNICODE).strip("_")
    return cleaned[:60] or "log"


async def send_lines(
    query: CallbackQuery,
    lines: list[str],
    header: str,
    filename: str,
    force_file: bool = False,
    caption: Optional[str] = None,
) -> None:
    """Довгий текст надсилаємо файлом, короткий — повідомленням.

    force_file=True — завжди файлом: лог по конкретному користувачу зручніше
    зберегти й погортати, ніж вишукувати серед повідомлень у чаті.
    """
    text = "\n".join(lines)
    if force_file or len(text) > TELEGRAM_TEXT_LIMIT or len(lines) > 200:
        tmp = DATA_DIR / filename
        try:
            await asyncio.to_thread(tmp.write_text, text, encoding="utf-8")
            with open(tmp, "rb") as fh:
                await query.message.reply_document(document=fh, filename=tmp.name, caption=caption or header)
        finally:
            try:
                tmp.unlink()
            except OSError:
                pass
        return
    await query.message.reply_text(f"{header}\n\n{text}")


# --------------------------------------------------------------------------- #
# Підтвердження адміном (review)
#
# Записи звичайних користувачів не пишуться в Excel одразу — створюється
# запис у _pending_reviews, а адміну летить лише легке сповіщення (без
# деталей і кнопок; повторно не турбуємо, поки черга не спорожніє — див.
# _notified_admins). Самі дії (підтвердити/відхилити/відредагувати) — тільки
# через кнопку меню «⏳ Очікують підтвердження». Перша відповідь виграє (під
# локом). Записи самих адмінів (is_admin) зберігаються без цього кроку —
# див. confirm_entry / process_edit_values.
# --------------------------------------------------------------------------- #

_pending_reviews: dict[str, dict] = {}

# Черга запитів дублюється на диск: інакше будь-який перезапуск процесу
# (оновлення, ребут, OOM) тихо з'їдав би всі непідтверджені записи — люди їх
# уже надіслали, а адмін їх більше не побачив би.
PENDING_FILE = DATA_DIR / "pending_reviews.json"


def _save_pending_reviews_sync() -> None:
    """Перезапис усієї черги. Файл крихітний (десятки записів), тож простіше
    писати цілком, ніж вести інкрементальний журнал. Запис через тимчасовий
    файл + os.replace — щоб падіння посеред запису не лишило огризок."""
    tmp_path = PENDING_FILE.with_suffix(".json.tmp")
    try:
        with open(tmp_path, "w", encoding="utf-8") as fh:
            json.dump(list(_pending_reviews.values()), fh, ensure_ascii=False, indent=2)
        os.replace(tmp_path, PENDING_FILE)
    except OSError as error:
        logger.warning("Cannot save pending reviews: %s", error)


def _load_pending_reviews_sync() -> int:
    """Піднімає чергу з диска при старті. Повертає кількість відновлених запитів."""
    if not PENDING_FILE.exists():
        return 0
    try:
        with open(PENDING_FILE, "r", encoding="utf-8") as fh:
            saved = json.load(fh)
    except (OSError, json.JSONDecodeError) as error:
        logger.warning("Cannot read pending reviews, starting with empty queue: %s", error)
        return 0

    restored = 0
    for review in saved if isinstance(saved, list) else []:
        if not isinstance(review, dict) or review.get("status") != "pending":
            continue
        if not review.get("id") or not review.get("user_id"):
            continue
        previous = review.get("previous")
        if isinstance(previous, list) and len(previous) == 2:
            review["previous"] = (int(previous[0]), int(previous[1]))
        _pending_reviews[review["id"]] = review
        restored += 1
    return restored


async def persist_pending_reviews() -> None:
    await asyncio.to_thread(_save_pending_reviews_sync)


def review_counts(review: dict) -> tuple[int, ...]:
    """Лічильники запиту. Розуміє і старі запити (big/small), які могли
    лишитись у pending_reviews.json з версії до появи winkel/dekel."""
    if "counts" in review:
        return normalize_counts(review["counts"])
    return normalize_counts({"box_177": review.get("big", 0), "box_161": review.get("small", 0)})


def _new_review_id() -> str:
    return uuid.uuid4().hex[:8]


def _user_label(user) -> str:
    username = getattr(user, "username", None)
    if username:
        return f"@{username}"
    full_name = getattr(user, "full_name", None)
    return full_name or str(getattr(user, "id", "?"))


async def _review_author_label(user) -> str:
    """Ім'я та прізвище, вписані при реєстрації — якщо їх ще немає, запасний варіант з Telegram."""
    registered_name = await get_registered_name(user.id)
    return registered_name or _user_label(user)


def _pending_review_count() -> int:
    return sum(1 for review in _pending_reviews.values() if review["status"] == "pending")


def _pending_reviews_sorted() -> list[dict]:
    """Усі запити зі статусом pending, у порядку надходження (найстаріші перші)."""
    reviews = [review for review in _pending_reviews.values() if review["status"] == "pending"]
    reviews.sort(key=lambda review: review.get("created_at", 0))
    return reviews


def _pending_reviews_by_user() -> dict[int, list[dict]]:
    """Групує запити по користувачу, зберігаючи хронологічний порядок усередині групи."""
    grouped: dict[int, list[dict]] = {}
    for review in _pending_reviews_sorted():
        grouped.setdefault(review["user_id"], []).append(review)
    return grouped


def _user_pending_reviews(user_id: int) -> list[dict]:
    """Власні запити цього користувача, що ще очікують підтвердження адміном."""
    return [review for review in _pending_reviews_sorted() if review["user_id"] == user_id]


def _user_pending_count(user_id: int) -> int:
    return len(_user_pending_reviews(user_id))


def _reviews_word(count: int) -> str:
    if count % 10 == 1 and count % 100 != 11:
        return "запит"
    if 2 <= count % 10 <= 4 and not (12 <= count % 100 <= 14):
        return "запити"
    return "запитів"


def _format_pending_users_list(grouped: dict[int, list[dict]]) -> str:
    if not grouped:
        return "⏳ Очікують підтвердження: немає запитів."
    total = sum(len(reviews) for reviews in grouped.values())
    return (
        f"⏳ Очікують підтвердження: {total} {_reviews_word(total)} "
        f"від {len(grouped)} користувача(ів).\n\nОбери, кого перевірити:"
    )


def _format_my_pending_reviews(reviews: list[dict]) -> str:
    """Список власних запитів користувача, що ще чекають на рішення адміна —
    без кнопок дій, це лише статус."""
    if not reviews:
        return "⏳ У тебе немає записів, що очікують підтвердження."
    lines = [f"⏳ Очікують підтвердження адміністратора: {len(reviews)} {_reviews_word(len(reviews))}.", ""]
    for review in reviews:
        if review["action"] == "delete":
            values = "видалити запис"
        else:
            counts = review_counts(review)
            values = f"{counts_summary(counts)} — {format_money(day_total(counts))} €"
        lines.append(f"• {review['date_text']}: {values}")
    return "\n".join(lines)


def build_my_pending_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[InlineKeyboardButton("🏠 Меню", callback_data="close_entry")]])


def build_pending_users_keyboard(grouped: dict[int, list[dict]]) -> InlineKeyboardMarkup:
    rows = []
    for uid, reviews in grouped.items():
        label = reviews[0]["user_label"]
        rows.append(
            [InlineKeyboardButton(f"👤 {label} ({len(reviews)})", callback_data=f"review_user_{uid}")]
        )
    rows.append([InlineKeyboardButton("🏠 Меню", callback_data="close_entry")])
    return InlineKeyboardMarkup(rows)


async def _next_view_after_resolution(review: dict, admin_id: int) -> tuple[str, InlineKeyboardMarkup]:
    """Що показати замість щойно обробленого запиту: наступний запит цього ж
    користувача (щоб адмін підтверджував декілька дат поспіль, по черзі),
    інакше — згорнутий список користувачів, інакше — що все оброблено."""
    remaining_same_user = [r for r in _pending_reviews_sorted() if r["user_id"] == review["user_id"]]
    if remaining_same_user:
        next_review = remaining_same_user[0]
        text = _format_review_text(next_review)
        if len(remaining_same_user) > 1:
            text += f"\n\n(Ще {len(remaining_same_user) - 1} після цього для {next_review['user_label']})"
        return text, build_review_keyboard(next_review["id"], include_back=True)

    grouped = _pending_reviews_by_user()
    if grouped:
        return _format_pending_users_list(grouped), build_pending_users_keyboard(grouped)
    return "⏳ Усі запити оброблено.", build_main_menu(user_id=admin_id)


async def _refresh_after_action(query: CallbackQuery, admin_id: int, review: dict) -> None:
    """Після підтвердження/відхилення перемальовує екран перегляду на наступний
    запит цього ж користувача (або на список користувачів, якщо більше нема
    що обробляти)."""
    if query.message is None:
        return
    text, keyboard = await _next_view_after_resolution(review, admin_id)
    await safe_edit(query, text, reply_markup=keyboard)


def build_review_keyboard(review_id: str, include_back: bool = False) -> InlineKeyboardMarkup:
    rows = [
        [
            InlineKeyboardButton("✅ Підтвердити", callback_data=f"review_approve_{review_id}"),
            InlineKeyboardButton("❌ Відхилити", callback_data=f"review_reject_{review_id}"),
        ],
        [InlineKeyboardButton("✏️ Редагувати", callback_data=f"review_edit_{review_id}")],
    ]
    if include_back:
        rows.append([InlineKeyboardButton("⬅️ До списку користувачів", callback_data="pending_reviews")])
    return InlineKeyboardMarkup(rows)


def _format_review_text(review: dict, status_line: Optional[str] = None) -> str:
    action_label = {
        "add": "🆕 Новий запис",
        "edit": "✏️ Зміна запису",
        "delete": "🗑 Видалення запису",
    }[review["action"]]
    lines = [
        action_label,
        f"Користувач: {review['user_label']} (ID {review['user_id']})",
        f"Дата: {review['date_text']}",
    ]
    previous = review.get("previous")
    if previous is not None:
        lines.append(f"Було: {counts_summary(previous)}")
    if review["action"] == "delete":
        lines.append("Дія: видалити запис за цей день")
    else:
        counts = review_counts(review)
        label = "Стало" if previous is not None else "Значення"
        lines.append(f"{label}: {counts_summary(counts)}")
        lines.append(f"Сума за день: {format_money(day_total(counts))} €")
    if status_line:
        lines.append("")
        lines.append(status_line)
    return "\n".join(lines)


# Кому вже пінганули про наявність нових запитів — щоб не засипати адміна
# повідомленнями, поки він не розгребе чергу. Скидається, щойно черга
# спорожніє (див. _reset_admin_notifications_if_empty), і наступний новий
# запит знову надішле сповіщення.
_notified_admins: set[int] = set()


async def _notify_admins_new_review(context: ContextTypes.DEFAULT_TYPE, review: dict) -> None:
    """Легке сповіщення без деталей і без кнопок: «є новий запит, дивись
    вкладку». Деталі та дії — тільки через «⏳ Очікують підтвердження» в
    меню. Поки в адміна лишається бодай один необроблений запит, повторно
    не турбуємо його новими такими повідомленнями."""
    text = "🔔 Новий запит на підтвердження.\nПеревір вкладку «⏳ Очікують підтвердження» в меню."
    for admin_id in ADMIN_IDS:
        if admin_id in _notified_admins:
            continue
        try:
            await context.bot.send_message(admin_id, text)
            _notified_admins.add(admin_id)
        except Exception as error:  # noqa: BLE001 — збій сповіщення не має ламати основний потік
            logger.warning("Could not notify admin %s about review %s: %s", admin_id, review["id"], error)


def _reset_admin_notifications_if_empty() -> None:
    """Коли черга запитів спорожніла — знімаємо позначку «вже сповіщений»,
    щоб наступний новий запит знову підняв адмінів."""
    if _pending_review_count() == 0:
        _notified_admins.clear()


# Нагадування про запити, які висять надто довго. Перше сповіщення про новий
# запит навмисно одноразове (щоб не спамити), тож без цього циклу забутий
# запит може лежати тижнями — людина чекає зарплату, адмін навіть не знає.
STALE_REVIEW_HOURS = float(_setting("STALE_REVIEW_HOURS", 24))
STALE_REMINDER_EVERY_HOURS = float(_setting("STALE_REMINDER_EVERY_HOURS", 24))
_STALE_CHECK_INTERVAL_SECONDS = 1800  # як часто перевіряти чергу (не як часто писати)

_last_stale_reminder: float = 0.0


def _stale_reviews(now: float) -> list[dict]:
    """Запити, що чекають довше за STALE_REVIEW_HOURS, найстаріші перші."""
    limit = STALE_REVIEW_HOURS * 3600
    return [
        review
        for review in _pending_reviews_sorted()
        if now - float(review.get("created_at", now)) >= limit
    ]


def _format_stale_reminder(reviews: list[dict], now: float) -> str:
    oldest = reviews[0]
    waiting_hours = int((now - float(oldest.get("created_at", now))) // 3600)
    count = len(reviews)
    return (
        f"⏰ Нагадування: {count} {_reviews_word(count)} без відповіді "
        f"понад {int(STALE_REVIEW_HOURS)} год.\n"
        f"Найдовше — {oldest['user_label']} за {oldest['date_text']} ({waiting_hours} год).\n"
        "Відкрий «⏳ Очікують підтвердження» в меню."
    )


async def _send_stale_reminder(bot) -> None:
    """Одна перевірка черги. Винесено окремо, щоб було що тестувати без циклу."""
    global _last_stale_reminder
    now = datetime.now(BERLIN_TZ).timestamp()
    reviews = _stale_reviews(now)
    if not reviews:
        return
    if now - _last_stale_reminder < STALE_REMINDER_EVERY_HOURS * 3600:
        return

    text = _format_stale_reminder(reviews, now)
    delivered = False
    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(admin_id, text)
            delivered = True
        except Exception as error:  # noqa: BLE001 — недоступний адмін не має зупиняти решту
            logger.warning("Could not send stale-review reminder to admin %s: %s", admin_id, error)
    if delivered:
        _last_stale_reminder = now
        logger.info("Нагадав адмінам про %d застряглих запитів", len(reviews))


async def _stale_reviews_watcher(application: Application) -> None:
    """Фоновий цикл. Свідомо не використовує JobQueue: вона тягне за собою
    APScheduler (python-telegram-bot[job-queue]), а тут вистачає sleep."""
    while True:
        try:
            await asyncio.sleep(_STALE_CHECK_INTERVAL_SECONDS)
            await _send_stale_reminder(application.bot)
        except asyncio.CancelledError:
            raise
        except Exception:  # noqa: BLE001 — цикл має пережити будь-який збій
            logger.exception("Stale-review watcher iteration failed")


async def _notify_user_result(context: ContextTypes.DEFAULT_TYPE, review: dict, resolution: str) -> None:
    """resolution: 'approved' | 'rejected' | 'edited'."""
    user_id = review["user_id"]
    date_text = review["date_text"]

    if resolution == "rejected":
        text = f"❌ Ваш запис за {date_text} відхилено адміністратором. Дані не збережено."
    elif review["action"] == "delete":
        prefix = "🗑 Ваш запис" if resolution == "approved" else "✏️ Адміністратор видалив ваш запис"
        text = f"{prefix} за {date_text} видалено."
    else:
        counts = review_counts(review)
        prefix = "✅ Ваш запис" if resolution == "approved" else "✏️ Адміністратор відредагував ваш запис"
        details = "\n".join(
            f"{CATEGORY_LABELS[key]}: {count}"
            for key, count in zip(CATEGORY_KEYS, counts)
            if count
        )
        text = (
            f"{prefix} за {date_text} підтверджено:\n"
            f"{details}\n"
            f"Сума за день: {format_money(day_total(counts))} €"
        )

    try:
        await context.bot.send_message(
            user_id, text, reply_markup=build_main_menu(user_id=user_id, month=review["month"])
        )
    except Exception as error:  # noqa: BLE001 — користувач міг заблокувати бота
        logger.warning("Could not notify user %s about review %s result: %s", user_id, review["id"], error)


# --------------------------------------------------------------------------- #
# Хендлери: введення дня
# --------------------------------------------------------------------------- #

DATE_HINT = "Введи дату: 15-08-2026\nМожна також: 15.08.2026 або 15,08,2026"
MENU_HINT = "Головне меню — обери дію:"
# Дата приймається тільки після «➕ Внести дані» / «📝 Редагувати дані»: інакше
# випадково надіслане в чат число з головного екрана мовчки заводило б запис.
NEED_ENTRY_BUTTON_HINT = (
    "Щоб внести дані, спочатку натисни «➕ Внести дані».\n"
    "Щоб виправити вже внесений день — «📝 Редагувати дані»."
)

WAITING_APPROVAL_TEXT = (
    "⏳ Заявку надіслано адміністратору.\n"
    "Щойно доступ підтвердять, бот напише — тоді натисни /start."
)


# --------------------------------------------------------------------------- #
# Екран вибору категорій
#
# Людина за день може робити кілька видів виробу поспіль (почала з коробок,
# перейшла на winkel), тому спершу відзначаємо галочками, що саме було, а вже
# потім бот питає кількість тільки по відзначених. Так у більшості випадків
# лишається два кроки, як раніше, а не шість зайвих питань.
# --------------------------------------------------------------------------- #


def _selected_keys(context: ContextTypes.DEFAULT_TYPE) -> list[str]:
    return [key for key in CATEGORY_KEYS if key in context.user_data.get("selected", [])]


def _entry_counts(context: ContextTypes.DEFAULT_TYPE) -> tuple[int, ...]:
    stored = context.user_data.get("counts", {})
    selected = set(_selected_keys(context))
    return normalize_counts({key: stored.get(key, 0) for key in selected})


def build_categories_keyboard(context: ContextTypes.DEFAULT_TYPE, allow_delete: bool = False) -> InlineKeyboardMarkup:
    selected = set(_selected_keys(context))
    counts = context.user_data.get("counts", {})
    rows = []
    for index in range(0, CATEGORY_COUNT, 2):
        row = []
        for key in CATEGORY_KEYS[index:index + 2]:
            mark = "✅" if key in selected else "▫️"
            count = counts.get(key, 0)
            suffix = f" ({count})" if key in selected and count else ""
            row.append(
                InlineKeyboardButton(f"{mark} {CATEGORY_LABELS[key]}{suffix}", callback_data=f"cat_{key}")
            )
        rows.append(row)
    rows.append([InlineKeyboardButton("▶️ Далі", callback_data="cat_next")])
    if allow_delete:
        rows.append([InlineKeyboardButton("🗑 Видалити запис за цей день", callback_data="cat_delete")])
    # Назад веде туди, звідки прийшли: у своєму записі — до вибору дати,
    # у правці чужого запиту — назад у чергу підтверджень.
    back_to = "pending_reviews" if context.user_data.get("flow") == "review" else "back_to_date"
    rows.append(
        [
            InlineKeyboardButton("⬅️ Назад", callback_data=back_to),
            InlineKeyboardButton("🏠 Меню", callback_data="close_entry"),
        ]
    )
    return InlineKeyboardMarkup(rows)


def categories_prompt(context: ContextTypes.DEFAULT_TYPE) -> str:
    date_text = context.user_data.get("date_text", "")
    who = context.user_data.get("review_user_label")
    header = f"Редагування запису {date_text}" + (f" для {who}" if who else "")
    if context.user_data.get("flow") != "review":
        header = f"Дата: {date_text}"
    lines = [header, "", "Що робив цього дня? Познач усе, що було:"]
    if context.user_data.get("previous"):
        lines.append("")
        lines.append(f"Зараз записано: {counts_summary(context.user_data['previous'])}")
    return "\n".join(lines)


async def show_categories(update_or_query, context: ContextTypes.DEFAULT_TYPE, edit: bool = False) -> int:
    allow_delete = context.user_data.get("previous") is not None and context.user_data.get("flow") != "review"
    keyboard = build_categories_keyboard(context, allow_delete=allow_delete)
    text = categories_prompt(context)
    if edit:
        await safe_edit(update_or_query, text, reply_markup=keyboard)
    else:
        await send_menu_reply(update_or_query, context, text, reply_markup=keyboard)
    return WAIT_CATEGORIES


async def ask_next_count(update_or_query, context: ContextTypes.DEFAULT_TYPE, edit: bool = False) -> int:
    """Питає кількість по наступній невідповіданій категорії; коли всі
    пройдено — показує екран перевірки."""
    queue = context.user_data.get("queue", [])
    if not queue:
        return await show_confirmation(update_or_query, context, edit=edit)

    key = queue[0]
    known = context.user_data.get("counts", {}).get(key)
    previous = context.user_data.get("previous")
    lines = [f"{CATEGORY_LABELS[key]} — скільки?"]
    if known:
        lines.append(f"(зараз {known} — введи нове число або те саме)")
    elif previous is not None:
        was = normalize_counts(previous)[CATEGORY_KEYS.index(key)]
        if was:
            lines.append(f"(було {was})")
    text = "\n".join(lines)
    keyboard = InlineKeyboardMarkup(
        [[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_categories")]]
    )
    if edit:
        await safe_edit(update_or_query, text, reply_markup=keyboard)
    else:
        await send_menu_reply(update_or_query, context, text, reply_markup=keyboard)
    return WAIT_COUNT


def _confirmation_text(context: ContextTypes.DEFAULT_TYPE) -> str:
    counts = _entry_counts(context)
    date_text = context.user_data.get("date_text", "")
    lines = ["Перевірка:", f"Дата: {date_text}"]
    who = context.user_data.get("review_user_label")
    if who:
        lines.append(f"Користувач: {who}")
    previous = context.user_data.get("previous")
    if previous is not None:
        lines.append(f"Було: {counts_summary(previous)}")
    lines.append("")
    lines += counts_lines(counts)
    lines.append(f"Сума за день: {format_money(day_total(counts))} €")
    lines.append("")
    lines.append("Все вірно?")
    return "\n".join(lines)


async def show_confirmation(update_or_query, context: ContextTypes.DEFAULT_TYPE, edit: bool = False) -> int:
    keyboard = InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton("💾 Зберегти", callback_data="save_entry"),
                InlineKeyboardButton("⬅️ Назад", callback_data="edit_entry"),
            ]
        ]
    )
    text = _confirmation_text(context)
    if edit:
        await safe_edit(update_or_query, text, reply_markup=keyboard)
    else:
        await send_menu_reply(update_or_query, context, text, reply_markup=keyboard)
    return WAIT_CONFIRM


async def on_category_button(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    data = query.data
    user_id = query.from_user.id

    if data == "cat_next":
        selected = _selected_keys(context)
        if not selected:
            await ack(query, "Познач хоча б одну категорію")
            return WAIT_CATEGORIES
        await ack(query)
        context.user_data["queue"] = [
            key for key in selected if not context.user_data.get("counts", {}).get(key)
        ] or []
        if not context.user_data["queue"]:
            return await show_confirmation(query, context, edit=True)
        return await ask_next_count(query, context, edit=True)

    if data == "cat_delete":
        await ack(query)
        context.user_data["counts"] = {}
        context.user_data["selected"] = []
        return await confirm_and_store(query, context, delete=True)

    key = data[len("cat_"):]
    if key not in CATEGORY_LABELS:
        await ack(query)
        return WAIT_CATEGORIES

    selected = set(_selected_keys(context))
    counts = dict(context.user_data.get("counts", {}))
    if key in selected:
        selected.discard(key)
        counts.pop(key, None)
    else:
        selected.add(key)
    context.user_data["selected"] = [k for k in CATEGORY_KEYS if k in selected]
    context.user_data["counts"] = counts
    await ack(query)
    return await show_categories(query, context, edit=True)


async def process_count(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await user_allowed(update.message.from_user.id):
        await send_menu_reply(update, context, WAITING_APPROVAL_TEXT)
        return ConversationHandler.END
    try:
        value = int(update.message.text.strip())
    except ValueError:
        await send_menu_reply(update, context, "Введи коректне число.")
        return WAIT_COUNT
    if value < 0:
        await send_menu_reply(update, context, "Кількість не може бути від'ємною.")
        return WAIT_COUNT

    queue = context.user_data.get("queue", [])
    if not queue:
        return await show_confirmation(update, context)

    key = queue.pop(0)
    counts = dict(context.user_data.get("counts", {}))
    counts[key] = value
    context.user_data["counts"] = counts
    context.user_data["queue"] = queue
    return await ask_next_count(update, context)


# --------------------------------------------------------------------------- #
# Реєстрація, доступ і дата
# --------------------------------------------------------------------------- #


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()
    user = update.message.from_user
    await register_user(user)
    logger.info("%s — відкрив бота (@%s)", user_tag(user.id), getattr(user, "username", None) or "—")

    registered_name = await get_registered_name(user.id)
    if not registered_name:
        await send_menu_reply(update, context,
            "Вітаю! Перш ніж почати, напиши, будь ласка, своє ім'я та прізвище "
            "(наприклад: Іван Петренко) — адміністратор бачитиме його в запитах на підтвердження."
        )
        return WAIT_REGISTER_NAME

    if not await user_allowed(user.id):
        await request_access(context, user.id, registered_name)
        await send_menu_reply(update, context, WAITING_APPROVAL_TEXT)
        return ConversationHandler.END

    await send_menu_reply(update, context, MENU_HINT, reply_markup=build_main_menu(user_id=user.id))
    return WAIT_DATE


async def process_registration_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    full_name = re.sub(r"\s+", " ", update.message.text.strip())[:100]
    has_letters = re.search(r"[^\W\d_]", full_name, re.UNICODE)
    if len(full_name) < 3 or not has_letters:
        await send_menu_reply(update, context,
            "Введи, будь ласка, ім'я та прізвище словами, наприклад: Іван Петренко"
        )
        return WAIT_REGISTER_NAME

    user = update.message.from_user
    await save_registered_name(user.id, full_name)
    logger.info("%s — представився як «%s»", user_tag(user.id, full_name), full_name)

    if not await user_allowed(user.id):
        await request_access(context, user.id, full_name)
        await send_menu_reply(update, context, f"Дякую, {full_name}!\n\n{WAITING_APPROVAL_TEXT}")
        return ConversationHandler.END

    await send_menu_reply(update, context,
        f"Дякую, {full_name}! Тепер можна вносити дані.\n\n{MENU_HINT}",
        reply_markup=build_main_menu(user_id=user.id),
    )
    return WAIT_DATE


async def open_day(target, context: ContextTypes.DEFAULT_TYPE, user_id: int, selected, edit_mode: bool, edit: bool = False) -> int:
    """Готує контекст під обрану дату і показує екран категорій. Спільне для
    вводу дати текстом і для кнопки «Сьогодні»."""
    month = selected.strftime("%Y-%m")
    context.user_data["month"] = month
    context.user_data["day"] = selected.day
    context.user_data["date_text"] = selected.strftime("%d-%m-%Y")
    context.user_data["flow"] = "entry"

    record = await read_day(user_id, month, selected.day)
    if edit_mode and record is None:
        text = f"Для {selected.strftime('%d-%m-%Y')} записів не знайдено."
        keyboard = build_main_menu(user_id=user_id, month=month)
        if edit:
            await safe_edit(target, text, reply_markup=keyboard)
        else:
            await send_menu_reply(target, context, text, reply_markup=keyboard)
        return WAIT_DATE

    context.user_data["previous"] = record
    if record is not None:
        # Уже внесений день відкриваємо з тими категоріями, що в ньому є —
        # людині лишається виправити цифру або дотиснути ще одну категорію.
        counts = normalize_counts(record)
        context.user_data["selected"] = [key for key, count in zip(CATEGORY_KEYS, counts) if count]
        context.user_data["counts"] = {key: count for key, count in zip(CATEGORY_KEYS, counts) if count}
    else:
        context.user_data["selected"] = []
        context.user_data["counts"] = {}
    return await show_categories(target, context, edit=edit)


async def process_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await user_allowed(update.message.from_user.id):
        await send_menu_reply(update, context, WAITING_APPROVAL_TEXT)
        return ConversationHandler.END

    if not context.user_data.get("awaiting_date"):
        await send_menu_reply(update, context,
            NEED_ENTRY_BUTTON_HINT,
            reply_markup=build_main_menu(user_id=update.message.from_user.id),
        )
        return WAIT_DATE

    raw_date = update.message.text.strip()
    if not DATE_PATTERN.match(raw_date):
        await send_menu_reply(update, context, "Невірний формат. Введи: 15-08-2026")
        return WAIT_DATE

    parsed = parse_date_input(raw_date)
    if parsed is None:
        await send_menu_reply(update, context, "Такої дати не існує. Введи: 15-08-2026")
        return WAIT_DATE

    selected = parsed.date()
    today = today_in_berlin()
    if selected > today:
        await send_menu_reply(update, context,
            f"Не можна вносити дані за майбутні дати. Сьогодні: {today.strftime('%d-%m-%Y')}."
        )
        return WAIT_DATE

    user_id = update.message.from_user.id
    await register_user(update.message.from_user)
    edit_mode = context.user_data.pop("edit_mode", False)
    context.user_data.pop("awaiting_date", None)
    return await open_day(update, context, user_id, selected, edit_mode)


# --------------------------------------------------------------------------- #
# Збереження дня (спільне для звичайного вводу і для правки адміном)
# --------------------------------------------------------------------------- #


async def confirm_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await ack(query)

    if query.data == "edit_entry":
        context.user_data["queue"] = []
        return await show_categories(query, context, edit=True)

    return await confirm_and_store(query, context)


async def confirm_and_store(query: CallbackQuery, context: ContextTypes.DEFAULT_TYPE, delete: bool = False) -> int:
    """Записує день або створює запит на підтвердження — залежно від того, хто
    вносить (адмін пише одразу) і чи це правка запиту в черзі."""
    user_id = query.from_user.id
    counts = EMPTY_COUNTS if delete else _entry_counts(context)

    # Правка чужого запиту йде своїм шляхом: там немає «свого» місяця й дня —
    # усе потрібне лежить у самому запиті.
    if context.user_data.get("flow") == "review":
        return await finish_review_edit(query, context, counts)

    month = context.user_data.get("month")
    day = context.user_data.get("day")
    date_text = context.user_data.get("date_text") or (f"{month}-{day:02d}" if month and day else "")

    if month is None or day is None:
        context.user_data.clear()
        await safe_edit(query, f"Дані втрачено. {MENU_HINT}", reply_markup=build_main_menu(user_id=user_id))
        return WAIT_DATE

    previous = context.user_data.get("previous")

    if not has_any_count(counts):
        if previous is None:
            context.user_data.clear()
            await safe_edit(
                query,
                "Нічого не збережено: усі кількості дорівнюють 0.",
                reply_markup=build_main_menu(user_id=user_id, month=month),
            )
            return WAIT_NEXT_ACTION
        action = "delete"
    else:
        action = "edit" if previous is not None else "add"

    await register_user(query.from_user)

    if is_admin(user_id):
        if action == "delete":
            removed, remaining = await delete_day(user_id, month, day)
            context.user_data.clear()
            if not removed:
                text = f"Запис за {date_text} не знайдено — нічого не змінено."
            elif remaining == 0:
                text = (
                    f"🗑 Запис {date_text} видалено.\n"
                    f"За {month} більше немає записів — місяць прибрано з меню."
                )
            else:
                monthly_total = await read_month_total(user_id, month)
                text = (
                    f"🗑 Запис {date_text} видалено.\n"
                    f"Днів у місяці лишилось: {remaining}\n"
                    f"Разом за місяць: {format_money(monthly_total)} €"
                )
            await safe_edit(query, text, reply_markup=build_main_menu(user_id=user_id, month=month))
            return WAIT_NEXT_ACTION

        total = await save_day(user_id, month, day, counts)
        monthly_total = await read_month_total(user_id, month)
        context.user_data.clear()
        lines = ["✅ Збережено" if action == "add" else "✏️ Оновлено", date_text]
        lines += counts_lines(counts)
        lines.append(f"Сума за день: {format_money(total)} €")
        lines.append(f"Разом за місяць: {format_money(monthly_total)} €")
        await safe_edit(
            query, "\n".join(lines), reply_markup=build_main_menu(user_id=user_id, month=month)
        )
        return WAIT_NEXT_ACTION

    review = {
        "id": _new_review_id(),
        "user_id": user_id,
        "user_label": await _review_author_label(query.from_user),
        "month": month,
        "day": day,
        "date_text": date_text,
        "action": action,
        "counts": list(counts),
        "previous": list(previous) if previous is not None else None,
        "status": "pending",
        "created_at": datetime.now(BERLIN_TZ).timestamp(),
    }
    _pending_reviews[review["id"]] = review
    await persist_pending_reviews()
    context.user_data.clear()

    if action == "delete":
        text = f"⏳ Запит на видалення запису за {date_text} надіслано адміністратору."
    else:
        lines = ["⏳ Надіслано адміністратору на підтвердження", date_text]
        lines += [f"{CATEGORY_LABELS[key]}: {count}" for key, count in zip(CATEGORY_KEYS, counts) if count]
        lines.append(f"Сума за день: {format_money(day_total(counts))} €")
        text = "\n".join(lines)

    await safe_edit(query, text, reply_markup=build_main_menu(user_id=user_id, month=month))
    await _notify_admins_new_review(context, review)
    logger.info(
        "%s — надіслав на підтвердження %s (%s): %s [запит %s]",
        user_tag(user_id), date_text, _ACTION_WORDS[action], counts_summary(counts), review["id"],
    )
    return WAIT_NEXT_ACTION


# --------------------------------------------------------------------------- #
# Хендлери: кнопки
# --------------------------------------------------------------------------- #


async def on_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обгортка: гарантує рівно одну відповідь на натискання. Змістовні
    повідомлення («даних ще немає») дає сам обробник, а якщо він промовчав —
    закриваємо натискання порожньою відповіддю, щоб кнопка не «крутилась»."""
    query = update.callback_query
    try:
        return await _dispatch_callback(update, context)
    finally:
        await ack(query)


async def _show_access_list(query: CallbackQuery, waiting: list[tuple[int, str]]) -> int:
    rows = []
    for uid, name in waiting:
        rows.append([InlineKeyboardButton(f"👤 {name} (ID {uid})", callback_data="noop")])
        rows.append(
            [
                InlineKeyboardButton("✅ Дозволити", callback_data=f"access_ok_{uid}"),
                InlineKeyboardButton("⛔ Відхилити", callback_data=f"access_no_{uid}"),
            ]
        )
    rows.append([InlineKeyboardButton("🏠 Меню", callback_data="close_entry")])
    await safe_edit(
        query, f"🔐 Заявки на доступ: {len(waiting)}\n\nКого пускаємо?", reply_markup=InlineKeyboardMarkup(rows)
    )
    return WAIT_NEXT_ACTION


async def _dispatch_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    data = query.data
    user_id = query.from_user.id
    admin = is_admin(user_id)
    await register_user(query.from_user)

    # Доступ дається адміном поіменно: чужі люди, які знайшли бота пошуком,
    # не мають ні вносити дані, ні смикати чергу підтверджень.
    if not admin and not await user_allowed(user_id):
        await ack(query, "Доступ ще не підтверджено адміністратором", alert=True)
        return WAIT_NEXT_ACTION

    if data == "noop":  # кнопка-підпис у списку заявок
        return WAIT_NEXT_ACTION

    # --- заявки на доступ (тільки адмін) ---------------------------------- #
    if data == "access_requests":
        if not admin:
            await ack(query, "Доступ заборонено", alert=True)
            return WAIT_NEXT_ACTION
        waiting = await pending_access()
        if not waiting:
            await ack(query, "Нових заявок немає")
            return WAIT_NEXT_ACTION
        return await _show_access_list(query, waiting)

    match = RE_ACCESS_OK.match(data) or RE_ACCESS_NO.match(data)
    if match:
        if not admin:
            await ack(query, "Доступ заборонено", alert=True)
            return WAIT_NEXT_ACTION
        target = int(match.group(1))
        granted = data.startswith("access_ok_")
        await set_access(target, granted)
        logger.info(
            "%s — %s доступ для %s",
            admin_tag(user_id), "надав" if granted else "відхилив", user_tag(target),
        )
        try:
            await context.bot.send_message(
                target,
                "✅ Доступ надано. Натисни /start, щоб почати вносити дані."
                if granted
                else "⛔ Адміністратор відхилив заявку на доступ.",
            )
        except Exception as error:  # noqa: BLE001 — користувач міг заблокувати бота
            logger.warning("Could not notify user %s about access decision: %s", target, error)

        waiting = await pending_access()
        if waiting:
            return await _show_access_list(query, waiting)
        await safe_edit(
            query,
            "✅ Доступ надано." if granted else "⛔ Заявку відхилено.",
            reply_markup=build_main_menu(user_id=user_id),
        )
        return WAIT_NEXT_ACTION

    # --- звичайні дії ---------------------------------------------------- #
    if data == "add_more":
        edit_mode = context.user_data.get("edit_mode", False)
        context.user_data.clear()
        context.user_data["edit_mode"] = edit_mode
        context.user_data["awaiting_date"] = True
        await safe_edit(
            query,
            f"За який день? {DATE_HINT}",
            reply_markup=build_date_keyboard(user_id=user_id),
        )
        return WAIT_DATE

    if data == "date_today":
        selected = today_in_berlin()
        edit_mode = context.user_data.pop("edit_mode", False)
        context.user_data.clear()  # заразом знімає awaiting_date
        return await open_day(query, context, user_id, selected, edit_mode, edit=True)

    if data == "edit_day_menu":
        context.user_data.clear()
        context.user_data["edit_mode"] = True
        context.user_data["awaiting_date"] = True
        await safe_edit(
            query,
            f"Який день редагуємо?\n{DATE_HINT}",
            reply_markup=build_date_keyboard(user_id=user_id),
        )
        return WAIT_DATE

    if data == "close_entry":
        context.user_data.clear()
        await safe_edit(query, MENU_HINT, reply_markup=build_main_menu(user_id=user_id))
        return WAIT_DATE

    if data == "back_to_date":
        # Назад з екрана категорій: питаємо дату знову, режим (новий запис чи
        # правка) зберігаємо.
        edit_mode = context.user_data.get("edit_mode", False) or context.user_data.get("previous") is not None
        context.user_data.clear()
        context.user_data["edit_mode"] = edit_mode
        context.user_data["awaiting_date"] = True
        await safe_edit(
            query,
            f"{'Який день редагуємо?' if edit_mode else 'За який день?'}\n{DATE_HINT}",
            reply_markup=build_date_keyboard(user_id=user_id),
        )
        return WAIT_DATE

    if data == "back_to_categories":
        if not context.user_data.get("month"):
            context.user_data["awaiting_date"] = True
            await safe_edit(query, f"За який день? {DATE_HINT}", reply_markup=build_date_keyboard(user_id=user_id))
            return WAIT_DATE
        context.user_data["queue"] = []
        return await show_categories(query, context, edit=True)

    if data == "show_stats":
        return await show_statistics(query, context)

    if data == "show_month_total":
        target = context.user_data.get("admin_view_user") if admin else user_id
        months = await list_months(None if (admin and target is None) else target)
        if not months:
            await ack(query, "Даних ще немає — спочатку додай день ➕")
            return WAIT_NEXT_ACTION
        await safe_edit(
            query,
            "Вибери місяць:",
            reply_markup=build_month_selection_keyboard(months, back_callback="show_stats"),
        )
        return WAIT_MONTH_INPUT

    match = RE_MONTH_TOTAL.match(data)
    if match:
        return await show_month_total(query, context, match.group(1))

    match = RE_DOWNLOAD.match(data)
    if match:
        return await send_month_files(query, context, match.group(1))

    if data == "my_pending_reviews":
        reviews = _user_pending_reviews(user_id)
        await safe_edit(query, _format_my_pending_reviews(reviews), reply_markup=build_my_pending_keyboard())
        return WAIT_NEXT_ACTION

    # --- підтвердження записів адміном ------------------------------------ #
    if data == "pending_reviews":
        if not admin:
            await ack(query, "Доступ заборонено", alert=True)
            return WAIT_NEXT_ACTION
        grouped = _pending_reviews_by_user()
        await safe_edit(query, _format_pending_users_list(grouped), reply_markup=build_pending_users_keyboard(grouped))
        return WAIT_NEXT_ACTION

    match = RE_REVIEW_USER.match(data)
    if match:
        return await handle_review_open_user(query, context, int(match.group(1)))

    match = RE_REVIEW_APPROVE.match(data)
    if match:
        return await handle_review_approve(query, context, match.group(1))

    match = RE_REVIEW_REJECT.match(data)
    if match:
        return await handle_review_reject(query, context, match.group(1))

    match = RE_REVIEW_EDIT.match(data)
    if match:
        return await handle_review_edit_start(query, context, match.group(1))

    # --- адмінські дії ---------------------------------------------------- #
    if data.startswith("admin_") and not admin:
        await ack(query, "Доступ заборонено", alert=True)
        logger.warning("%s — спроба адмінської дії без прав: %s", user_tag(user_id), data)
        return WAIT_NEXT_ACTION

    if data == "admin_monthly_summary":
        months = await list_months(None)
        if not months:
            await ack(query, "Даних ще немає — спочатку додай день ➕")
            return WAIT_NEXT_ACTION
        await safe_edit(
            query,
            "За який місяць зробити зведену таблицю по всіх користувачах?",
            reply_markup=build_admin_summary_month_keyboard(months),
        )
        return WAIT_NEXT_ACTION

    match = RE_ADMIN_SUMMARY.match(data)
    if match:
        return await send_monthly_summary(query, match.group(1))

    if data in ("admin_users", "admin_log_by_user"):
        users = await known_users()
        if not users:
            await ack(query, "Користувачів ще немає")
            return WAIT_NEXT_ACTION
        prefix = "admin_user_" if data == "admin_users" else "admin_log_user_"
        rows = []
        for uid, (name, has_data) in users.items():
            label = f"{uid} — {name}" if name else str(uid)
            if not has_data:
                label += " · без даних"
            rows.append([InlineKeyboardButton(label, callback_data=f"{prefix}{uid}")])
        back_row = []
        if data == "admin_log_by_user":
            back_row.append(InlineKeyboardButton("⬅️ Назад", callback_data="admin_logs"))
        back_row.append(InlineKeyboardButton("🏠 Меню", callback_data="close_entry"))
        rows.append(back_row)
        title = "Оберіть користувача:" if data == "admin_users" else "Логи якого користувача показати?"
        await safe_edit(query, title, reply_markup=InlineKeyboardMarkup(rows))
        return WAIT_NEXT_ACTION

    match = RE_ADMIN_USER.match(data)
    if match:
        selected = int(match.group(1))
        context.user_data["admin_view_user"] = selected
        months = await list_months(selected)
        if not months:
            await ack(query, "У цього користувача немає даних")
            return WAIT_NEXT_ACTION
        await safe_edit(
            query,
            f"Оберіть місяць для користувача {selected}:",
            reply_markup=build_month_selection_keyboard(months, back_callback="admin_users"),
        )
        return WAIT_MONTH_INPUT

    if data == "admin_logs":
        keyboard = InlineKeyboardMarkup(
            [
                [InlineKeyboardButton("📄 Повний файл", callback_data="admin_log_file")],
                [InlineKeyboardButton("👤 По користувачу", callback_data="admin_log_by_user")],
                [InlineKeyboardButton("🏠 Меню", callback_data="close_entry")],
            ]
        )
        await safe_edit(query, "Журнал (логи): оберіть опцію:", reply_markup=keyboard)
        return WAIT_NEXT_ACTION

    if data == "admin_log_file":
        if not LOG_FILE.exists():
            await ack(query, "Файл лога не знайдено")
            return WAIT_NEXT_ACTION
        with open(LOG_FILE, "rb") as fh:
            await query.message.reply_document(document=fh, filename=LOG_FILE.name)
        logger.info("%s — вивантажив повний файл лога", admin_tag(user_id))
        return WAIT_NEXT_ACTION

    match = RE_ADMIN_LOG_USER.match(data)
    if match:
        # Проміжного меню («останні 100/500/повний файл») немає навмисно:
        # обрав людину — одразу отримав її повний лог файлом.
        selected = int(match.group(1))
        result = await filter_log(f"[USER:{selected}]")
        if not result:
            await ack(query, "Записів для цього користувача не знайдено")
            return WAIT_NEXT_ACTION
        name = await get_registered_name(selected) or str(selected)
        await send_lines(
            query,
            result,
            f"Повний лог — {name} ({len(result)} {_lines_word(len(result))})",
            f"Лог_{_safe_filename(name)}.txt",
            force_file=True,
        )
        logger.info("%s — вивантажив повний лог по %s", admin_tag(user_id), user_tag(selected))
        return WAIT_NEXT_ACTION

    return WAIT_NEXT_ACTION


async def handle_review_open_user(query: CallbackQuery, context: ContextTypes.DEFAULT_TYPE, target_user_id: int) -> int:
    """Показує перший (найстаріший) запит вибраного користувача з кнопками дій.
    Якщо в нього декілька дат — після кожної обробленої показується наступна,
    тож адмін підтверджує/відхиляє їх по черзі, не вертаючись щоразу у список."""
    if not is_admin(query.from_user.id):
        await ack(query, "Доступ заборонено", alert=True)
        return WAIT_NEXT_ACTION

    reviews = [r for r in _pending_reviews_sorted() if r["user_id"] == target_user_id]
    if not reviews:
        await ack(query, "У цього користувача більше немає запитів")
        grouped = _pending_reviews_by_user()
        await safe_edit(query, _format_pending_users_list(grouped), reply_markup=build_pending_users_keyboard(grouped))
        return WAIT_NEXT_ACTION

    review = reviews[0]
    text = _format_review_text(review)
    if len(reviews) > 1:
        text += f"\n\n(Ще {len(reviews) - 1} після цього для {review['user_label']})"
    await safe_edit(query, text, reply_markup=build_review_keyboard(review["id"], include_back=True))
    return WAIT_NEXT_ACTION


async def handle_review_approve(query: CallbackQuery, context: ContextTypes.DEFAULT_TYPE, review_id: str) -> int:
    admin_id = query.from_user.id
    if not is_admin(admin_id):
        await ack(query, "Доступ заборонено", alert=True)
        return WAIT_NEXT_ACTION

    async with _lock_for(f"review:{review_id}"):
        review = _pending_reviews.get(review_id)
        if review is None or review["status"] != "pending":
            await ack(query, "Цей запит вже оброблено або застарів.", alert=True)
            return WAIT_NEXT_ACTION
        review["status"] = "approved"
        review["resolved_by"] = admin_id
        del _pending_reviews[review_id]  # оброблений запит більше не тримаємо в пам'яті
    _locks.pop(f"review:{review_id}", None)
    await persist_pending_reviews()

    if review["action"] == "delete":
        await delete_day(review["user_id"], review["month"], review["day"])
    else:
        await save_day(review["user_id"], review["month"], review["day"], review_counts(review))

    _reset_admin_notifications_if_empty()
    await _refresh_after_action(query, admin_id, review)
    await _notify_user_result(context, review, "approved")
    logger.info(
        "%s — підтвердив %s за %s (%s) [запит %s]",
        admin_tag(admin_id), _ACTION_WORDS[review["action"]], review["date_text"],
        user_tag(review["user_id"], review.get("user_label")), review_id,
    )
    return WAIT_NEXT_ACTION


async def handle_review_reject(query: CallbackQuery, context: ContextTypes.DEFAULT_TYPE, review_id: str) -> int:
    admin_id = query.from_user.id
    if not is_admin(admin_id):
        await ack(query, "Доступ заборонено", alert=True)
        return WAIT_NEXT_ACTION

    async with _lock_for(f"review:{review_id}"):
        review = _pending_reviews.get(review_id)
        if review is None or review["status"] != "pending":
            await ack(query, "Цей запит вже оброблено або застарів.", alert=True)
            return WAIT_NEXT_ACTION
        review["status"] = "rejected"
        review["resolved_by"] = admin_id
        del _pending_reviews[review_id]
    _locks.pop(f"review:{review_id}", None)
    await persist_pending_reviews()

    _reset_admin_notifications_if_empty()
    await _refresh_after_action(query, admin_id, review)
    await _notify_user_result(context, review, "rejected")
    logger.info(
        "[ADMIN:%s] Rejected review %s for [USER:%s] %s",
        admin_id, review_id, review["user_id"], review["date_text"],
    )
    return WAIT_NEXT_ACTION


async def handle_review_edit_start(query: CallbackQuery, context: ContextTypes.DEFAULT_TYPE, review_id: str) -> int:
    """Адмін правит чужий запит тим самим екраном категорій, що й свій запис."""
    admin_id = query.from_user.id
    if not is_admin(admin_id):
        await ack(query, "Доступ заборонено", alert=True)
        return WAIT_NEXT_ACTION

    review = _pending_reviews.get(review_id)
    if review is None or review["status"] != "pending":
        await ack(query, "Цей запит вже оброблено або застарів.", alert=True)
        return WAIT_NEXT_ACTION

    counts = review_counts(review)
    context.user_data.clear()
    context.user_data["flow"] = "review"
    context.user_data["admin_edit_review_id"] = review_id
    context.user_data["date_text"] = review["date_text"]
    context.user_data["review_user_label"] = review["user_label"]
    context.user_data["previous"] = review.get("previous")
    context.user_data["selected"] = [key for key, count in zip(CATEGORY_KEYS, counts) if count]
    context.user_data["counts"] = {key: count for key, count in zip(CATEGORY_KEYS, counts) if count}
    if query.message is not None:
        context.user_data["admin_edit_origin_chat_id"] = query.message.chat_id
        context.user_data["admin_edit_origin_message_id"] = query.message.message_id
    return await show_categories(query, context, edit=True)


async def finish_review_edit(query: CallbackQuery, context: ContextTypes.DEFAULT_TYPE, counts) -> int:
    """Застосовує рішення адміна до запиту в черзі. Якщо цифри не змінилися —
    для користувача це звичайне підтвердження, а не «адмін відредагував»."""
    admin_id = query.from_user.id
    review_id = context.user_data.get("admin_edit_review_id")
    origin_chat_id = context.user_data.get("admin_edit_origin_chat_id")
    origin_message_id = context.user_data.get("admin_edit_origin_message_id")
    context.user_data.clear()

    values = normalize_counts(counts)

    async with _lock_for(f"review:{review_id}"):
        review = _pending_reviews.get(review_id)
        if review is None or review["status"] != "pending":
            await safe_edit(query, "Цей запит вже оброблено іншим адміном.")
            return WAIT_NEXT_ACTION
        unchanged = values == review_counts(review)
        review["status"] = "approved" if unchanged else "edited"
        review["resolved_by"] = admin_id
        if not unchanged:
            review["counts"] = list(values)
            review.pop("big", None)
            review.pop("small", None)
            review["action"] = "delete" if not has_any_count(values) else "edit"
        del _pending_reviews[review_id]
    _locks.pop(f"review:{review_id}", None)
    await persist_pending_reviews()

    if review["action"] == "delete":
        await delete_day(review["user_id"], review["month"], review["day"])
    else:
        await save_day(review["user_id"], review["month"], review["day"], review_counts(review))

    _reset_admin_notifications_if_empty()
    if origin_chat_id is not None and origin_message_id is not None:
        text, keyboard = await _next_view_after_resolution(review, admin_id)
        try:
            await context.bot.edit_message_text(
                chat_id=origin_chat_id, message_id=origin_message_id, text=text, reply_markup=keyboard
            )
        except Exception as error:  # noqa: BLE001 — це лише зручність навігації, не критично
            logger.debug("Could not advance origin review screen: %s", error)

    await _notify_user_result(context, review, "approved" if unchanged else "edited")
    logger.info(
        "%s — %s запис за %s (%s): %s [запит %s]",
        admin_tag(admin_id),
        "підтвердив без змін" if unchanged else "змінив і зберіг",
        review["date_text"], user_tag(review["user_id"], review.get("user_label")),
        counts_summary(values), review_id,
    )
    await safe_edit(
        query,
        "✅ Підтверджено без змін, користувача повідомлено."
        if unchanged
        else "Збережено, користувача повідомлено.",
    )
    return WAIT_NEXT_ACTION


def _month_title(month: str) -> str:
    months_ua = [
        "січень", "лютий", "березень", "квітень", "травень", "червень",
        "липень", "серпень", "вересень", "жовтень", "листопад", "грудень",
    ]
    try:
        parsed = datetime.strptime(month, "%Y-%m")
        return f"{months_ua[parsed.month - 1]} {parsed.year}"
    except (ValueError, IndexError):
        return month


async def build_statistics_text(user_id: int, month: str, owner_label: Optional[str] = None) -> str:
    stats = await month_statistics(user_id, month)
    header = f"📊 Статистика за {_month_title(month)}"
    if owner_label:
        header += f" — {owner_label}"

    if not stats["days"]:
        return f"{header}\n\nЗа цей місяць ще немає жодного робочого дня."

    lines = [
        header,
        "",
        f"💶 Заробіток: {format_money(stats['total'])} €",
        f"📆 Робочих днів: {stats['days']}",
        f"📈 У середньому за день: {format_money(stats['average'])} €",
    ]
    if stats["best_day"]:
        lines.append(
            f"🥇 Найкращий день: {_human_date(stats['best_day'])} — {format_money(stats['best_money'])} €"
        )
    made = counts_summary(stats["counts"])
    if made != "—":
        lines += ["", f"Зроблено: {made}"]
    return "\n".join(lines)


async def show_statistics(query: CallbackQuery, context: ContextTypes.DEFAULT_TYPE, month: Optional[str] = None) -> int:
    """Екран «Статистика»: поточний місяць підтягується сам, без вибору зі списку."""
    user_id = query.from_user.id
    admin_target = context.user_data.pop("admin_view_user", None) if is_admin(user_id) else None
    target_id = admin_target if admin_target is not None else user_id
    month = month or current_month()

    owner_label = None
    if admin_target is not None:
        user_map = await load_user_map()
        entry = user_map.get(str(target_id), {})
        owner_label = (entry.get("registered_name") or entry.get("telegram_label") or "").strip() or str(target_id)

    text = await build_statistics_text(target_id, month, owner_label)
    context.user_data["download_user"] = target_id

    if await month_has_data(user_file(target_id, month)):
        keyboard = build_month_result_keyboard(month)
    else:
        # Поточний місяць порожній: пропонуємо або внести день, або глянути
        # інший місяць — але тільки якщо ті інші місяці взагалі є.
        rows = [[InlineKeyboardButton("➕ Внести дані", callback_data="add_more")]]
        if await list_months(target_id):
            rows.append([InlineKeyboardButton("🗓 Інший місяць", callback_data="show_month_total")])
        rows.append([InlineKeyboardButton("🏠 Меню", callback_data="close_entry")])
        keyboard = InlineKeyboardMarkup(rows)

    await safe_edit(query, text, reply_markup=keyboard)
    return WAIT_NEXT_ACTION


async def show_month_total(query: CallbackQuery, context: ContextTypes.DEFAULT_TYPE, month: str) -> int:
    user_id = query.from_user.id
    admin_target = context.user_data.pop("admin_view_user", None) if is_admin(user_id) else None
    target_id = admin_target if admin_target is not None else user_id

    path = user_file(target_id, month)
    if not await month_has_data(path):
        await ack(query, "За цей місяць немає записів")
        return WAIT_NEXT_ACTION

    owner_label = None
    if admin_target is not None:
        user_map = await load_user_map()
        entry = user_map.get(str(target_id), {})
        owner_label = (entry.get("registered_name") or entry.get("telegram_label") or "").strip() or str(target_id)

    context.user_data["download_user"] = target_id
    await safe_edit(
        query,
        await build_statistics_text(target_id, month, owner_label),
        reply_markup=build_month_result_keyboard(month, back_callback="show_month_total"),
    )
    return WAIT_NEXT_ACTION


async def send_month_files(query: CallbackQuery, context: ContextTypes.DEFAULT_TYPE, month: str) -> int:
    user_id = query.from_user.id
    target_id = context.user_data.pop("download_user", user_id)
    if target_id != user_id and not is_admin(user_id):
        target_id = user_id

    path = user_file(target_id, month)
    if not path.exists():
        await ack(query, "Файл не знайдено")
        return WAIT_NEXT_ACTION

    user_map = await load_user_map()
    entry = user_map.get(str(target_id), {})
    display_name = ""
    if isinstance(entry, dict):
        display_name = (entry.get("registered_name") or entry.get("telegram_label") or "").strip()

    report_path = None
    try:
        report_path = await build_user_report(target_id, month, display_name)
        if report_path is None:
            await ack(query, "За цей місяць немає записів")
            return WAIT_NEXT_ACTION
        with open(report_path, "rb") as fh:
            await query.message.reply_document(
                document=fh, filename=f"Заробіток_{month}.xlsx", caption=f"Звіт за {month}"
            )
        logger.info("%s — отримав звіт за %s", user_tag(user_id), month)
    except (OSError, BadRequest) as error:
        logger.exception("%s — не вдалося надіслати звіт за %s: %s", user_tag(user_id), month, error)
        await query.message.reply_text("Не вдалося надіслати файл. Спробуй ще раз.")
    finally:
        if report_path is not None:
            try:
                report_path.unlink()
            except OSError:
                pass
    return WAIT_NEXT_ACTION


async def send_monthly_summary(query: CallbackQuery, month: str) -> int:
    """Зведена таблиця за місяць: один рядок на користувача (коробки кожного
    типу, разом і сума), відсортовано за алфавітом — для швидкого огляду адміном."""
    admin_id = query.from_user.id
    path = await build_monthly_summary(month)
    if path is None:
        await ack(query, f"За {month} ще немає жодних даних")
        return WAIT_NEXT_ACTION

    try:
        with open(path, "rb") as fh:
            await query.message.reply_document(document=fh, filename=f"Зведення_{month}.xlsx")
        logger.info("%s — отримав зведення за %s", admin_tag(admin_id), month)
    except (OSError, BadRequest) as error:
        logger.exception("%s — не вдалося надіслати зведення за %s: %s", admin_tag(admin_id), month, error)
        await query.message.reply_text("Не вдалося надіслати таблицю. Спробуй ще раз.")
    finally:
        try:
            path.unlink()
        except OSError:
            pass
    return WAIT_NEXT_ACTION


async def process_month_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    month = update.message.text.strip()
    if not MONTH_PATTERN.match(month):
        await send_menu_reply(update, context, "Невірний формат місяця. Введи YYYY-MM, наприклад 2026-08.")
        return WAIT_MONTH_INPUT

    user_id = update.message.from_user.id
    admin_target = context.user_data.pop("admin_view_user", None) if is_admin(user_id) else None

    if is_admin(user_id) and admin_target is None:
        total, names = await read_all_totals(month)
        if not names:
            await send_menu_reply(update, context, 
                f"За місяць {month} ще немає записів.",
                reply_markup=build_main_menu(user_id=user_id, month=month),
            )
            return WAIT_NEXT_ACTION
        await send_menu_reply(update, context, 
            f"Усі користувачі за {month}: {format_money(total)} €\nФайли: {', '.join(names)}",
            reply_markup=build_main_menu(user_id=user_id, month=month),
        )
        return WAIT_NEXT_ACTION

    target_id = admin_target if admin_target is not None else user_id
    path = user_file(target_id, month)
    if not await month_has_data(path):
        await send_menu_reply(update, context, 
            f"За місяць {month} ще немає записів.",
            reply_markup=build_main_menu(user_id=user_id, month=month),
        )
        return WAIT_NEXT_ACTION

    context.user_data["download_user"] = target_id
    await send_menu_reply(update, context,
        await build_statistics_text(target_id, month),
        reply_markup=build_month_result_keyboard(month),
    )
    return WAIT_NEXT_ACTION


# --------------------------------------------------------------------------- #
# Команди
# --------------------------------------------------------------------------- #


async def cmd_total(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    args = context.args
    month = args[0].strip() if args else current_month()
    if not MONTH_PATTERN.match(month):
        await send_menu_reply(update, context, "Формат: /total 2026-08")
        return

    user_id = update.message.from_user.id
    if is_admin(user_id):
        total, names = await read_all_totals(month)
        if not names:
            await send_menu_reply(update, context, f"За місяць {month} ще немає записів.")
            return
        await send_menu_reply(update, context, 
            f"Усі користувачі за {month}: {format_money(total)} €\nФайли: {', '.join(names)}"
        )
        return

    path = user_file(user_id, month)
    if not await month_has_data(path):
        await send_menu_reply(update, context, f"За місяць {month} ще немає записів.")
        return
    total = await read_month_total(user_id, month)
    await send_menu_reply(update, context, f"{month}: {format_money(total)} €\nФайл: {path.name}")


async def cmd_logtail(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """/logtail [N|file] — тільки для адмінів."""
    user_id = update.message.from_user.id
    if not is_admin(user_id):
        await send_menu_reply(update, context, "Доступ заборонено.")
        logger.warning("%s — спроба /logtail без прав", user_tag(user_id))
        return

    mode = context.args[0].strip().lower() if context.args else "100"

    if mode == "file":
        if not LOG_FILE.exists():
            await send_menu_reply(update, context, "Файл лога не знайдено.")
            return
        with open(LOG_FILE, "rb") as fh:
            await update.message.reply_document(document=fh, filename=LOG_FILE.name)
        logger.info("%s — вивантажив повний файл лога (/logtail)", admin_tag(user_id))
        return

    try:
        lines = min(int(mode), 2000)
    except ValueError:
        lines = 100

    result = await tail_log(lines)
    if not result:
        await send_menu_reply(update, context, "Лог порожній.")
        return

    text = "\n".join(result)
    if len(text) > TELEGRAM_TEXT_LIMIT or len(result) > 200:
        tmp = DATA_DIR / f"log_tail_{user_id}.txt"
        try:
            await asyncio.to_thread(tmp.write_text, text, encoding="utf-8")
            with open(tmp, "rb") as fh:
                await update.message.reply_document(document=fh, filename=tmp.name)
        finally:
            try:
                tmp.unlink()
            except OSError:
                pass
    else:
        await send_menu_reply(update, context, f"Останні {len(result)} рядків лога:\n\n{text}")
    logger.info("%s — переглянув лог (%d рядків)", admin_tag(user_id), len(result))


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()
    await send_menu_reply(update, context, "Скасовано. Для нового запису натисни /start.")
    return ConversationHandler.END


async def on_error(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    error = context.error

    if isinstance(error, Conflict):
        logger.error(
            "Conflict: з цим токеном уже працює інший екземпляр бота. "
            "Зупини другий процес (systemctl stop salarybot) або візьми окремий тестовий токен."
        )
        return

    if isinstance(error, NetworkError) and not isinstance(error, BadRequest):
        logger.warning("Мережева помилка: %s", error)
        return

    logger.error("Unhandled exception", exc_info=error)
    if isinstance(update, Update) and update.effective_message:
        try:
            await update.effective_message.reply_text(
                "Сталася помилка. Спробуй ще раз або натисни /start."
            )
        except Exception:  # noqa: BLE001 — не даємо помилці в обробнику помилок впасти далі
            logger.exception("Failed to notify user about error")


# --------------------------------------------------------------------------- #
# Точка входу
# --------------------------------------------------------------------------- #


async def _post_init(application: Application) -> None:
    """Налаштовує список команд, який Telegram показує в кнопці «Меню» біля
    поля вводу (те саме, що BotFather /setcommands). Адмінам показуємо ще й
    /logtail — окремою «областю видимості» лише для їхніх особистих чатів."""
    default_commands = [
        BotCommand("start", "Почати"),
        BotCommand("menu", "Головне меню"),
        BotCommand("total", "Сума за поточний місяць"),
        BotCommand("cancel", "Скасувати поточну дію"),
    ]
    try:
        await application.bot.set_my_commands(default_commands, scope=BotCommandScopeDefault())
    except Exception as error:  # noqa: BLE001 — відсутність меню команд не критична
        logger.warning("Could not set default command menu: %s", error)

    admin_commands = default_commands + [BotCommand("logtail", "Останні рядки логів (адмін)")]
    for admin_id in ADMIN_IDS:
        try:
            await application.bot.set_my_commands(admin_commands, scope=BotCommandScopeChat(chat_id=admin_id))
        except Exception as error:  # noqa: BLE001 — адмін міг ще жодного разу не писати боту
            logger.warning("Could not set admin command menu for %s: %s", admin_id, error)

    application.bot_data["stale_watcher"] = asyncio.create_task(_stale_reviews_watcher(application))
    logger.info(
        "Нагадування про застряглі запити: раз на %s год про ті, що старші за %s год",
        int(STALE_REMINDER_EVERY_HOURS), int(STALE_REVIEW_HOURS),
    )


async def _post_shutdown(application: Application) -> None:
    task = application.bot_data.get("stale_watcher")
    if task is not None:
        task.cancel()


class PerChatUpdateProcessor(BaseUpdateProcessor):
    """Апдейти різних чатів обробляються паралельно (до max_concurrent_updates
    одночасно) — щоб один активний користувач не змушував решту команди
    чекати в черзі. Апдейти ОДНОГО чату обробляються строго послідовно, щоб
    діалоговий стан того самого користувача не зіткнувся сам із собою —
    штатний concurrent_updates=True такого не гарантує (апдейти одного чату
    теж могли б піти паралельно)."""

    def __init__(self, max_concurrent_updates: int) -> None:
        super().__init__(max_concurrent_updates)
        self._chat_locks: dict[int, asyncio.Lock] = {}

    def _lock_for_chat(self, chat_id: int) -> asyncio.Lock:
        lock = self._chat_locks.get(chat_id)
        if lock is None:
            lock = asyncio.Lock()
            self._chat_locks[chat_id] = lock
        return lock

    async def do_process_update(self, update: object, coroutine) -> None:
        chat = getattr(update, "effective_chat", None)
        if chat is None:
            await coroutine
            return
        lock = self._lock_for_chat(chat.id)
        async with lock:
            await coroutine
        # Замок живе рівно стільки, скільки триває обробка: інакше словник ріс
        # би на кожен новий чат і ніколи не порожнів. Якщо на замку хтось чекає,
        # він лишається захопленим — і ми його не чіпаємо.
        if not lock.locked():
            self._chat_locks.pop(chat.id, None)

    async def initialize(self) -> None:
        pass

    async def shutdown(self) -> None:
        pass


def main() -> None:
    token = BOT_TOKEN or os.getenv("BOT_TOKEN")
    if not token or token == "YOUR_BOT_TOKEN_HERE":
        raise RuntimeError(
            "BOT_TOKEN не задано. Вкажи його в config.py або в змінній оточення BOT_TOKEN."
        )

    # Починаючи з Python 3.12 asyncio.get_event_loop() не створює цикл сам,
    # а в 3.14 кидає RuntimeError. PTB викликає його всередині run_polling(),
    # тому створюємо цикл явно.
    try:
        asyncio.get_event_loop()
    except RuntimeError:
        asyncio.set_event_loop(asyncio.new_event_loop())

    application = (
        Application.builder()
        .token(token)
        .concurrent_updates(PerChatUpdateProcessor(32))
        .post_init(_post_init)
        .post_shutdown(_post_shutdown)
        .build()
    )

    callback_handler = CallbackQueryHandler(on_callback, pattern=CALLBACK_PATTERN)

    conversation = ConversationHandler(
        entry_points=[CommandHandler("start", start), CommandHandler("menu", start)],
        states={
            WAIT_DATE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_date),
                callback_handler,
            ],
            WAIT_CATEGORIES: [
                CallbackQueryHandler(on_category_button, pattern=r"^cat_"),
                callback_handler,
            ],
            WAIT_COUNT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_count),
                CallbackQueryHandler(on_category_button, pattern=r"^cat_"),
                callback_handler,
            ],
            WAIT_CONFIRM: [
                CallbackQueryHandler(confirm_entry, pattern=r"^(save_entry|edit_entry)$"),
                callback_handler,
            ],
            WAIT_REGISTER_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_registration_name),
                callback_handler,
            ],
            WAIT_MONTH_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_month_input),
                callback_handler,
            ],
            WAIT_NEXT_ACTION: [callback_handler],
        },
        fallbacks=[
            CommandHandler("cancel", cancel),
            CommandHandler("start", start),
            CommandHandler("menu", start),
        ],
        allow_reentry=True,
    )

    application.add_handler(conversation)
    application.add_handler(CommandHandler("total", cmd_total))
    application.add_handler(CommandHandler("logtail", cmd_logtail))
    application.add_error_handler(on_error)

    removed = cleanup_empty_files()
    if removed:
        logger.info("Старт: прибрано порожніх файлів місяця — %d", removed)

    restored = _load_pending_reviews_sync()
    if restored:
        logger.info("Старт: відновлено запитів у черзі — %d", restored)

    logger.info("─" * 60)
    logger.info("Бот запущено. Версія: %s", BOT_VERSION)
    logger.info("Адміни: %s", ", ".join(str(a) for a in sorted(ADMIN_IDS)))
    logger.info("Категорії: %s", ", ".join(CATEGORY_LABELS[key] for key in CATEGORY_KEYS))
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()